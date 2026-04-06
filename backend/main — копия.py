import os
import uuid
import time
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import List, Optional

from fastapi import FastAPI, Depends, HTTPException, Query, UploadFile, File, status, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO

import models
import schemas
import database
from models import (
    Supplier, SupplierPriceFile, SupplierPriceItem,
    MaterialRequest, MaterialRequestItem
)

# ===== Инициализация =====
models.Base.metadata.create_all(bind=database.engine)

app = FastAPI(title="CRM Repair System", version="3.2.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

BASE_DIR = Path(__file__).parent
FRONTEND_DIR = BASE_DIR / "frontend"
UPLOADS_DIR = BASE_DIR / "uploads"
UPLOADS_DIR.mkdir(exist_ok=True)
UPLOAD_DIR = str(UPLOADS_DIR)

# Конфигурация
CONTRACTOR_MARGIN = float(os.getenv("CONTRACTOR_MARGIN", "0.5"))
MAX_FILE_SIZE = int(os.getenv("MAX_FILE_SIZE", str(50 * 1024 * 1024)))

if FRONTEND_DIR.exists():
    app.mount("/static", StaticFiles(directory=str(FRONTEND_DIR)), name="static")


# ===== Утилиты =====
def now_str():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def today_str():
    return date.today().isoformat()


def escape_like(s: str) -> str:
    return s.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")


def safe_float(val, default=0.0):
    try:
        return float(val or default)
    except (ValueError, TypeError):
        return default


def calc_object_financials_from_loaded(obj: models.Object) -> dict:
    debit = sum(p.amount for p in obj.payments if p.status == "paid")
    contractor = sum(w.total_price for w in obj.contractor_works if w.status in ("completed", "paid"))
    material = sum(m.total_price for m in obj.material_purchases if m.status == "delivered")
    credit = contractor + material
    profit = debit - credit
    margin = (profit / debit * 100) if debit > 0 else 0
    return {
        "total_debit": round(float(debit), 2),
        "contractor_expenses": round(float(contractor), 2),
        "material_expenses": round(float(material), 2),
        "total_credit": round(float(credit), 2),
        "profit": round(float(profit), 2),
        "margin_percent": round(margin, 1),
    }


def calc_object_financials_db(obj_id: int, db: Session) -> dict:
    debit = db.query(func.sum(models.Payment.amount)).filter(
        models.Payment.object_id == obj_id, models.Payment.status == "paid"
    ).scalar() or 0
    contractor = db.query(func.sum(models.ContractorWork.total_price)).filter(
        models.ContractorWork.object_id == obj_id,
        models.ContractorWork.status.in_(["completed", "paid"]),
    ).scalar() or 0
    material = db.query(func.sum(models.MaterialPurchase.total_price)).filter(
        models.MaterialPurchase.object_id == obj_id,
        models.MaterialPurchase.status == "delivered",
    ).scalar() or 0
    credit = contractor + material
    profit = debit - credit
    margin = (profit / debit * 100) if debit > 0 else 0
    return {
        "total_debit": round(float(debit), 2),
        "contractor_expenses": round(float(contractor), 2),
        "material_expenses": round(float(material), 2),
        "total_credit": round(float(credit), 2),
        "profit": round(float(profit), 2),
        "margin_percent": round(margin, 1),
    }


def _object_joinedload_options():
    return [
        joinedload(models.Object.estimates).joinedload(models.Estimate.items),
        joinedload(models.Object.payments),
        joinedload(models.Object.contractor_works).joinedload(models.ContractorWork.contractor),
        joinedload(models.Object.contractor_works).joinedload(models.ContractorWork.tools),
        joinedload(models.Object.material_purchases).joinedload(models.MaterialPurchase.material),
        joinedload(models.Object.documents),
        joinedload(models.Object.communications),
        joinedload(models.Object.tasks),
        joinedload(models.Object.contractor_estimates).joinedload(models.ContractorEstimate.items),
        joinedload(models.Object.contractor_estimates).joinedload(models.ContractorEstimate.contractor),
    ]


def parse_estimate_excel(file_path: str):
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        items = []
        header_row = None
        headers = {}
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=False), 1):
            row_values = [str(c.value).strip().lower() if c.value else "" for c in row]
            for col_idx, val in enumerate(row_values):
                if "наименован" in val or "работ" in val:
                    headers["name"] = col_idx
                    header_row = row_idx
                elif "ед" in val and ("изм" in val or "." in val):
                    headers["unit"] = col_idx
                elif "объём" in val or "объем" in val or "кол" in val:
                    headers["quantity"] = col_idx
                elif "цена" in val and "ед" in val:
                    headers["price"] = col_idx
                elif "сумм" in val or "стоим" in val:
                    headers["total"] = col_idx
            if "name" in headers:
                break
        if not header_row:
            headers = {"name": 0, "unit": 1, "quantity": 2, "price": 3, "total": 4}
            header_row = 1
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            row = list(row) + [None] * 10
            name_val = row[headers.get("name", 0)]
            if not name_val or str(name_val).strip() == "":
                continue
            unit = str(row[headers.get("unit", 1)] or "").strip()
            quantity = safe_float(row[headers.get("quantity", 2)])
            price = safe_float(row[headers.get("price", 3)])
            total = safe_float(row[headers.get("total", 4)])
            if total == 0 and quantity > 0 and price > 0:
                total = quantity * price
            items.append({
                "name": str(name_val).strip(),
                "unit": unit,
                "quantity": quantity,
                "price_per_unit": price,
                "total_price": total,
            })
        return items
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Ошибка парсинга Excel: {str(e)}")


# =====================
# RESPONSE BUILDERS
# =====================

def build_object_response(obj: models.Object, use_loaded_financials: bool = True) -> dict:
    fin = calc_object_financials_from_loaded(obj) if use_loaded_financials else {}
    return {
        "id": obj.id,
        "name": obj.name,
        "client_name": obj.client_name,
        "client_phone": obj.client_phone,
        "client_email": obj.client_email,
        "client_address": obj.client_address,
        "status": obj.status,
        "created_at": obj.created_at,
        "notes": obj.notes,
        "estimates": [
            {
                "id": e.id, "object_id": e.object_id, "name": e.name,
                "file_path": e.file_path, "created_at": e.created_at,
                "items": [
                    {"id": i.id, "estimate_id": i.estimate_id, "name": i.name,
                     "unit": i.unit, "quantity": i.quantity,
                     "price_per_unit": i.price_per_unit, "total_price": i.total_price}
                    for i in e.items
                ],
            }
            for e in obj.estimates
        ],
        "payments": [
            {"id": p.id, "object_id": p.object_id, "amount": p.amount,
             "status": p.status, "description": p.description, "date": p.date}
            for p in obj.payments
        ],
        "contractor_works": [build_work_response(w) for w in obj.contractor_works],
        "material_purchases": [
            {
                "id": m.id, "object_id": m.object_id, "material_id": m.material_id,
                "quantity": m.quantity, "total_price": m.total_price,
                "supplier": m.supplier, "date": m.date, "status": m.status, "notes": m.notes,
                "material": {
                    "id": m.material.id, "name": m.material.name,
                    "unit": m.material.unit, "price_per_unit": m.material.price_per_unit,
                    "description": m.material.description,
                } if m.material else None,
            }
            for m in obj.material_purchases
        ],
        "documents": [
            {"id": d.id, "object_id": d.object_id, "doc_type": d.doc_type,
             "name": d.name, "file_path": d.file_path, "created_at": d.created_at}
            for d in obj.documents
        ],
        "communications": [
            {"id": c.id, "object_id": c.object_id, "type": c.type,
             "description": c.description, "date": c.date}
            for c in obj.communications
        ],
        "tasks": [
            {"id": t.id, "object_id": t.object_id, "title": t.title,
             "description": t.description, "status": t.status,
             "deadline": t.deadline, "created_at": t.created_at}
            for t in obj.tasks
        ],
        "contractor_estimates": [
            build_contractor_estimate_response(e) for e in obj.contractor_estimates
        ],
        "financials": {
            "object_id": obj.id, "object_name": obj.name,
            "client_name": obj.client_name, **fin,
        },
    }


def build_work_response(w: models.ContractorWork) -> dict:
    return {
        "id": w.id, "object_id": w.object_id, "contractor_id": w.contractor_id,
        "estimate_item_id": w.estimate_item_id, "description": w.description,
        "unit": w.unit, "quantity": w.quantity, "price_per_unit": w.price_per_unit,
        "total_price": w.total_price, "advance": w.advance, "deadline": w.deadline,
        "status": w.status, "notes": w.notes, "created_at": w.created_at,
        "contractor": {
            "id": w.contractor.id, "name": w.contractor.name,
            "phone": w.contractor.phone, "specialization": w.contractor.specialization,
            "notes": w.contractor.notes,
        } if w.contractor else None,
        "tools": [
            {"id": t.id, "name": t.name, "serial_number": t.serial_number,
             "purchase_price": t.purchase_price, "purchase_date": t.purchase_date,
             "status": t.status, "contractor_id": t.contractor_id,
             "object_id": t.object_id, "notes": t.notes}
            for t in w.tools
        ],
    }


def build_contractor_estimate_response(e: models.ContractorEstimate) -> dict:
    contractor = e.contractor
    return {
        "id": e.id, "object_id": e.object_id, "contractor_id": e.contractor_id,
        "name": e.name, "status": e.status, "total_sum": e.total_sum,
        "created_at": e.created_at, "notes": e.notes,
        "items": [
            {"id": i.id, "estimate_id": i.estimate_id, "name": i.name,
             "unit": i.unit, "quantity": i.quantity,
             "price_per_unit": i.price_per_unit, "total_price": i.total_price}
            for i in e.items
        ],
        "contractor": {
            "id": contractor.id, "name": contractor.name,
            "phone": contractor.phone, "specialization": contractor.specialization,
        } if contractor else None,
    }


def build_purchase_response(p: models.MaterialPurchase) -> dict:
    mat = p.material
    return {
        "id": p.id, "object_id": p.object_id, "material_id": p.material_id,
        "quantity": p.quantity, "total_price": p.total_price,
        "supplier": p.supplier, "date": p.date, "status": p.status, "notes": p.notes,
        "material": {
            "id": mat.id, "name": mat.name, "unit": mat.unit,
            "price_per_unit": mat.price_per_unit, "description": mat.description,
        } if mat else None,
    }


# =====================
# ОБЪЕКТЫ
# =====================

@app.post("/objects/", response_model=schemas.ObjectFullResponse)
async def create_object(obj: schemas.ObjectCreate, db: Session = Depends(database.get_db)):
    db_obj = models.Object(**obj.model_dump(), created_at=now_str())
    db.add(db_obj)
    db.commit()
    db.refresh(db_obj)
    return build_object_response(db_obj)


@app.get("/objects/", response_model=List[schemas.ObjectFullResponse])
async def list_objects(
    status: Optional[str] = None,
    search: Optional[str] = None,
    db: Session = Depends(database.get_db),
):
    q = db.query(models.Object).options(*_object_joinedload_options())
    if status:
        q = q.filter(models.Object.status == status)
    if search:
        safe_search = escape_like(search)
        q = q.filter(
            models.Object.name.ilike(f"%{safe_search}%")
            | models.Object.client_name.ilike(f"%{safe_search}%")
        )
    objects = q.order_by(models.Object.id.desc()).all()
    return [build_object_response(o) for o in objects]


@app.get("/objects/{obj_id}", response_model=schemas.ObjectFullResponse)
async def get_object(obj_id: int, db: Session = Depends(database.get_db)):
    obj = db.query(models.Object).options(
        *_object_joinedload_options()
    ).filter(models.Object.id == obj_id).first()
    if not obj:
        raise HTTPException(404, "Объект не найден")
    return build_object_response(obj)


@app.put("/objects/{obj_id}", response_model=schemas.ObjectFullResponse)
async def update_object(obj_id: int, data: schemas.ObjectUpdate, db: Session = Depends(database.get_db)):
    obj = db.query(models.Object).options(
        *_object_joinedload_options()
    ).filter(models.Object.id == obj_id).first()
    if not obj:
        raise HTTPException(404, "Объект не найден")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(obj, k, v)
    db.commit()
    db.refresh(obj)
    return build_object_response(obj)


@app.delete("/objects/{obj_id}")
async def delete_object(obj_id: int, db: Session = Depends(database.get_db)):
    obj = db.query(models.Object).filter(models.Object.id == obj_id).first()
    if not obj:
        raise HTTPException(404, "Объект не найден")
    db.delete(obj)
    db.commit()
    return {"ok": True}


# =====================
# СМЕТЫ
# =====================

@app.post("/objects/{obj_id}/estimate/upload")
async def upload_estimate(obj_id: int, file: UploadFile = File(...), db: Session = Depends(database.get_db)):
    obj = db.query(models.Object).filter(models.Object.id == obj_id).first()
    if not obj:
        raise HTTPException(404, "Объект не найден")
    ext = Path(file.filename).suffix.lower()
    if ext not in (".xlsx", ".xls", ".pdf"):
        raise HTTPException(400, "Поддерживаются только Excel (.xlsx) и PDF файлы")
    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(413, f"Файл слишком большой. Максимум: {MAX_FILE_SIZE // (1024 * 1024)} MB")
    safe_name = f"estimate_{obj_id}_{uuid.uuid4().hex[:8]}{ext}"
    file_path = UPLOADS_DIR / safe_name
    with open(file_path, "wb") as f:
        f.write(content)
    estimate_count = db.query(models.Estimate).filter(models.Estimate.object_id == obj_id).count()
    estimate = models.Estimate(
        object_id=obj_id, name=f"Смета №{estimate_count + 1}",
        file_path=str(file_path), created_at=now_str(),
    )
    db.add(estimate)
    db.flush()
    items_count = 0
    if ext in (".xlsx", ".xls"):
        items_data = parse_estimate_excel(str(file_path))
        for item in items_data:
            db.add(models.EstimateItem(
                estimate_id=estimate.id, name=item["name"], unit=item["unit"],
                quantity=item["quantity"], price_per_unit=item["price_per_unit"],
                total_price=item["total_price"],
            ))
            items_count += 1
    db.commit()
    return {
        "estimate_id": estimate.id, "name": estimate.name,
        "items_count": items_count, "file_path": safe_name,
        "message": f"Смета загружена. Распознано позиций: {items_count}" if ext != ".pdf" else "PDF сохранён",
    }


@app.delete("/estimates/{estimate_id}")
async def delete_estimate(estimate_id: int, db: Session = Depends(database.get_db)):
    est = db.query(models.Estimate).filter(models.Estimate.id == estimate_id).first()
    if not est:
        raise HTTPException(404, "Смета не найдена")
    if est.file_path and os.path.exists(est.file_path):
        os.remove(est.file_path)
    db.delete(est)
    db.commit()
    return {"ok": True}


@app.get("/objects/{obj_id}/estimates")
async def get_object_estimates(obj_id: int, db: Session = Depends(database.get_db)):
    estimates = db.query(models.Estimate).options(
        joinedload(models.Estimate.items)
    ).filter(models.Estimate.object_id == obj_id).all()
    return [{
        "id": e.id, "name": e.name, "created_at": e.created_at,
        "items_count": len(e.items),
        "total_sum": sum(i.total_price for i in e.items),
        "items": [{"id": i.id, "name": i.name, "unit": i.unit,
                    "quantity": i.quantity, "price_per_unit": i.price_per_unit,
                    "total_price": i.total_price} for i in e.items],
    } for e in estimates]


@app.get("/objects/{obj_id}/estimate-items/search")
async def search_estimate_items(obj_id: int, q: str = "", db: Session = Depends(database.get_db)):
    safe_q = escape_like(q)
    items = db.query(models.EstimateItem).join(models.Estimate).filter(
        models.Estimate.object_id == obj_id,
        models.EstimateItem.name.ilike(f"%{safe_q}%"),
    ).limit(20).all()
    return [{
        "id": i.id, "estimate_id": i.estimate_id, "name": i.name,
        "unit": i.unit, "quantity": i.quantity,
        "price_per_unit": i.price_per_unit, "total_price": i.total_price,
        "contractor_price": round(i.total_price * CONTRACTOR_MARGIN, 2),
        "contractor_price_per_unit": round(i.price_per_unit * CONTRACTOR_MARGIN, 2),
    } for i in items]


# =====================
# ДОКУМЕНТЫ ОБЪЕКТА
# =====================

@app.post("/objects/{obj_id}/document/upload")
async def upload_object_document(
    obj_id: int, doc_type: str = "act",
    file: UploadFile = File(...), db: Session = Depends(database.get_db),
):
    obj = db.query(models.Object).filter(models.Object.id == obj_id).first()
    if not obj:
        raise HTTPException(404, "Объект не найден")
    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(413, f"Файл слишком большой. Максимум: {MAX_FILE_SIZE // (1024 * 1024)} MB")
    ext = Path(file.filename).suffix.lower()
    safe_name = f"doc_{obj_id}_{uuid.uuid4().hex[:8]}{ext}"
    file_path = UPLOADS_DIR / safe_name
    with open(file_path, "wb") as f:
        f.write(content)
    doc = models.ObjectDocument(
        object_id=obj_id, doc_type=doc_type, name=file.filename,
        file_path=str(file_path), created_at=now_str(),
    )
    db.add(doc)
    db.commit()
    db.refresh(doc)
    return {"id": doc.id, "name": doc.name, "doc_type": doc.doc_type, "file_path": safe_name}


@app.get("/uploads/{filename}")
async def get_uploaded_file(filename: str):
    file_path = (UPLOADS_DIR / filename).resolve()
    if not str(file_path).startswith(str(UPLOADS_DIR.resolve())):
        raise HTTPException(403, "Доступ запрещён")
    if not file_path.exists():
        raise HTTPException(404, "Файл не найден")
    return FileResponse(str(file_path))


# =====================
# ПЛАТЕЖИ
# =====================

@app.post("/payments/", response_model=schemas.PaymentResponse)
async def create_payment(data: schemas.PaymentCreate, db: Session = Depends(database.get_db)):
    payload = data.model_dump()
    if not payload.get("date"):
        payload["date"] = today_str()
    db_p = models.Payment(**payload)
    db.add(db_p)
    db.commit()
    db.refresh(db_p)
    return db_p


@app.get("/payments/")
async def list_payments(object_id: Optional[int] = None, db: Session = Depends(database.get_db)):
    q = db.query(models.Payment)
    if object_id:
        q = q.filter(models.Payment.object_id == object_id)
    return q.order_by(models.Payment.id.desc()).all()


@app.put("/payments/{pay_id}", response_model=schemas.PaymentResponse)
async def update_payment(pay_id: int, data: schemas.PaymentUpdate, db: Session = Depends(database.get_db)):
    p = db.query(models.Payment).filter(models.Payment.id == pay_id).first()
    if not p:
        raise HTTPException(404, "Платёж не найден")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(p, k, v)
    db.commit()
    db.refresh(p)
    return p


@app.delete("/payments/{pay_id}")
async def delete_payment(pay_id: int, db: Session = Depends(database.get_db)):
    p = db.query(models.Payment).filter(models.Payment.id == pay_id).first()
    if not p:
        raise HTTPException(404, "Платёж не найден")
    db.delete(p)
    db.commit()
    return {"ok": True}


# =====================
# ПОДРЯДЧИКИ
# =====================

@app.post("/contractors/", response_model=schemas.ContractorResponse)
async def create_contractor(data: schemas.ContractorCreate, db: Session = Depends(database.get_db)):
    c = models.Contractor(**data.model_dump())
    db.add(c)
    db.commit()
    db.refresh(c)
    return c


@app.get("/contractors/", response_model=List[schemas.ContractorResponse])
async def list_contractors(db: Session = Depends(database.get_db)):
    return db.query(models.Contractor).all()


@app.get("/contractors/{cid}", response_model=schemas.ContractorResponse)
async def get_contractor(cid: int, db: Session = Depends(database.get_db)):
    c = db.query(models.Contractor).filter(models.Contractor.id == cid).first()
    if not c:
        raise HTTPException(404, "Подрядчик не найден")
    return c


@app.put("/contractors/{cid}", response_model=schemas.ContractorResponse)
async def update_contractor(cid: int, data: schemas.ContractorUpdate, db: Session = Depends(database.get_db)):
    c = db.query(models.Contractor).filter(models.Contractor.id == cid).first()
    if not c:
        raise HTTPException(404, "Подрядчик не найден")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(c, k, v)
    db.commit()
    db.refresh(c)
    return c


@app.delete("/contractors/{cid}")
async def delete_contractor(cid: int, db: Session = Depends(database.get_db)):
    c = db.query(models.Contractor).filter(models.Contractor.id == cid).first()
    if not c:
        raise HTTPException(404, "Подрядчик не найден")
    db.delete(c)
    db.commit()
    return {"ok": True}


# =====================
# РАБОТЫ ПОДРЯДЧИКОВ
# =====================

@app.post("/contractor-works/")
async def create_work(data: schemas.ContractorWorkCreate, db: Session = Depends(database.get_db)):
    d = data.model_dump()
    tool_ids = d.pop("tool_ids", [])
    if d.get("estimate_item_id"):
        item = db.query(models.EstimateItem).filter(models.EstimateItem.id == d["estimate_item_id"]).first()
        if item:
            if not d.get("quantity"):
                d["quantity"] = item.quantity
            if not d.get("unit"):
                d["unit"] = item.unit
            if not d.get("price_per_unit"):
                d["price_per_unit"] = round(item.price_per_unit * CONTRACTOR_MARGIN, 2)
            if not d.get("total_price"):
                d["total_price"] = round(d["quantity"] * d["price_per_unit"], 2)
            if not d.get("description"):
                d["description"] = item.name
    if not d.get("description"):
        raise HTTPException(400, "Укажите описание работы")
    work = models.ContractorWork(**d, created_at=now_str())
    db.add(work)
    db.flush()
    if tool_ids:
        tools = db.query(models.Tool).filter(models.Tool.id.in_(tool_ids)).all()
        work.tools = tools
    db.commit()
    db.refresh(work)
    return build_work_response(work)


@app.get("/contractor-works/")
async def list_works(
    object_id: Optional[int] = None,
    contractor_id: Optional[int] = None,
    status: Optional[str] = None,
    db: Session = Depends(database.get_db),
):
    q = db.query(models.ContractorWork).options(
        joinedload(models.ContractorWork.contractor),
        joinedload(models.ContractorWork.tools),
    )
    if object_id:
        q = q.filter(models.ContractorWork.object_id == object_id)
    if contractor_id:
        q = q.filter(models.ContractorWork.contractor_id == contractor_id)
    if status:
        q = q.filter(models.ContractorWork.status == status)
    return [build_work_response(w) for w in q.order_by(models.ContractorWork.id.desc()).all()]


@app.put("/contractor-works/{work_id}")
async def update_work(work_id: int, data: schemas.ContractorWorkUpdate, db: Session = Depends(database.get_db)):
    w = db.query(models.ContractorWork).filter(models.ContractorWork.id == work_id).first()
    if not w:
        raise HTTPException(404, "Работа не найдена")
    update_data = data.model_dump(exclude_unset=True)
    tool_ids = update_data.pop("tool_ids", None)
    for k, v in update_data.items():
        setattr(w, k, v)
    if tool_ids is not None:
        w.tools = db.query(models.Tool).filter(models.Tool.id.in_(tool_ids)).all()
    db.commit()
    db.refresh(w)
    return build_work_response(w)


@app.delete("/contractor-works/{work_id}")
async def delete_work(work_id: int, db: Session = Depends(database.get_db)):
    w = db.query(models.ContractorWork).filter(models.ContractorWork.id == work_id).first()
    if not w:
        raise HTTPException(404, "Работа не найдена")
    db.delete(w)
    db.commit()
    return {"ok": True}


# =====================
# МАТЕРИАЛЫ
# =====================

@app.post("/materials/", response_model=schemas.MaterialResponse)
async def create_material(data: schemas.MaterialCreate, db: Session = Depends(database.get_db)):
    m = models.Material(**data.model_dump())
    db.add(m)
    db.commit()
    db.refresh(m)
    return m


@app.get("/materials/", response_model=List[schemas.MaterialResponse])
async def list_materials(db: Session = Depends(database.get_db)):
    return db.query(models.Material).all()


@app.put("/materials/{mid}", response_model=schemas.MaterialResponse)
async def update_material(mid: int, data: schemas.MaterialUpdate, db: Session = Depends(database.get_db)):
    m = db.query(models.Material).filter(models.Material.id == mid).first()
    if not m:
        raise HTTPException(404, "Материал не найден")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(m, k, v)
    db.commit()
    db.refresh(m)
    return m


@app.delete("/materials/{mid}")
async def delete_material(mid: int, db: Session = Depends(database.get_db)):
    m = db.query(models.Material).filter(models.Material.id == mid).first()
    if not m:
        raise HTTPException(404, "Материал не найден")
    db.delete(m)
    db.commit()
    return {"ok": True}


# =====================
# ЗАКУПКИ МАТЕРИАЛОВ
# =====================

@app.post("/material-purchases/")
async def create_purchase(data: schemas.MaterialPurchaseCreate, db: Session = Depends(database.get_db)):
    payload = data.model_dump()
    if not payload.get("date"):
        payload["date"] = today_str()
    p = models.MaterialPurchase(**payload)
    db.add(p)
    db.commit()
    db.refresh(p)
    _ = p.material
    return build_purchase_response(p)


@app.get("/material-purchases/")
async def list_purchases(object_id: Optional[int] = None, db: Session = Depends(database.get_db)):
    q = db.query(models.MaterialPurchase).options(joinedload(models.MaterialPurchase.material))
    if object_id:
        q = q.filter(models.MaterialPurchase.object_id == object_id)
    return [build_purchase_response(p) for p in q.order_by(models.MaterialPurchase.id.desc()).all()]


@app.put("/material-purchases/{pid}")
async def update_purchase(pid: int, data: schemas.MaterialPurchaseUpdate, db: Session = Depends(database.get_db)):
    p = db.query(models.MaterialPurchase).options(
        joinedload(models.MaterialPurchase.material)
    ).filter(models.MaterialPurchase.id == pid).first()
    if not p:
        raise HTTPException(404, "Закупка не найдена")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(p, k, v)
    db.commit()
    db.refresh(p)
    return build_purchase_response(p)


@app.delete("/material-purchases/{pid}")
async def delete_purchase(pid: int, db: Session = Depends(database.get_db)):
    p = db.query(models.MaterialPurchase).filter(models.MaterialPurchase.id == pid).first()
    if not p:
        raise HTTPException(404, "Закупка не найдена")
    db.delete(p)
    db.commit()
    return {"ok": True}


# =====================
# ИНСТРУМЕНТЫ
# =====================

@app.post("/tools/", response_model=schemas.ToolResponse)
async def create_tool(data: schemas.ToolCreate, db: Session = Depends(database.get_db)):
    t = models.Tool(**data.model_dump())
    db.add(t)
    db.commit()
    db.refresh(t)
    return t


@app.get("/tools/", response_model=List[schemas.ToolResponse])
async def list_tools(
    contractor_id: Optional[int] = None,
    status: Optional[str] = None,
    db: Session = Depends(database.get_db),
):
    q = db.query(models.Tool)
    if contractor_id:
        q = q.filter(models.Tool.contractor_id == contractor_id)
    if status:
        q = q.filter(models.Tool.status == status)
    return q.all()


@app.get("/tools/{tid}", response_model=schemas.ToolResponse)
async def get_tool(tid: int, db: Session = Depends(database.get_db)):
    t = db.query(models.Tool).filter(models.Tool.id == tid).first()
    if not t:
        raise HTTPException(404, "Инструмент не найден")
    return t


@app.put("/tools/{tid}", response_model=schemas.ToolResponse)
async def update_tool(tid: int, data: schemas.ToolUpdate, db: Session = Depends(database.get_db)):
    t = db.query(models.Tool).filter(models.Tool.id == tid).first()
    if not t:
        raise HTTPException(404, "Инструмент не найден")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(t, k, v)
    db.commit()
    db.refresh(t)
    return t


@app.delete("/tools/{tid}")
async def delete_tool(tid: int, db: Session = Depends(database.get_db)):
    t = db.query(models.Tool).filter(models.Tool.id == tid).first()
    if not t:
        raise HTTPException(404, "Инструмент не найден")
    db.delete(t)
    db.commit()
    return {"ok": True}


# =====================
# ЗАДАЧИ
# =====================

@app.post("/tasks/")
async def create_task(data: schemas.TaskCreate, db: Session = Depends(database.get_db)):
    t = models.Task(**data.model_dump(), created_at=now_str())
    db.add(t)
    db.commit()
    db.refresh(t)
    return t


@app.get("/tasks/")
async def list_tasks(object_id: Optional[int] = None, db: Session = Depends(database.get_db)):
    q = db.query(models.Task)
    if object_id:
        q = q.filter(models.Task.object_id == object_id)
    return q.all()


@app.put("/tasks/{tid}")
async def update_task(tid: int, data: schemas.TaskUpdate, db: Session = Depends(database.get_db)):
    t = db.query(models.Task).filter(models.Task.id == tid).first()
    if not t:
        raise HTTPException(404, "Задача не найдена")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(t, k, v)
    db.commit()
    db.refresh(t)
    return t


@app.delete("/tasks/{tid}")
async def delete_task(tid: int, db: Session = Depends(database.get_db)):
    t = db.query(models.Task).filter(models.Task.id == tid).first()
    if not t:
        raise HTTPException(404, "Задача не найдена")
    db.delete(t)
    db.commit()
    return {"ok": True}


# =====================
# КОММУНИКАЦИИ
# =====================

@app.post("/communications/")
async def create_comm(data: schemas.CommunicationCreate, db: Session = Depends(database.get_db)):
    c = models.Communication(**data.model_dump())
    db.add(c)
    db.commit()
    db.refresh(c)
    return c


@app.get("/communications/")
async def list_comms(object_id: Optional[int] = None, db: Session = Depends(database.get_db)):
    q = db.query(models.Communication)
    if object_id:
        q = q.filter(models.Communication.object_id == object_id)
    return q.order_by(models.Communication.date.desc()).all()


# =====================
# ОБЩИЕ РАСХОДЫ
# =====================

@app.post("/general-expenses/")
async def create_general_expense(data: schemas.GeneralExpenseCreate, db: Session = Depends(database.get_db)):
    payload = data.model_dump()
    if not payload.get("date"):
        payload["date"] = today_str()
    e = models.GeneralExpense(**payload)
    db.add(e)
    db.commit()
    db.refresh(e)
    return e


@app.get("/general-expenses/")
async def list_general_expenses(db: Session = Depends(database.get_db)):
    return db.query(models.GeneralExpense).order_by(models.GeneralExpense.id.desc()).all()


@app.put("/general-expenses/{eid}")
async def update_general_expense(eid: int, data: schemas.GeneralExpenseUpdate, db: Session = Depends(database.get_db)):
    e = db.query(models.GeneralExpense).filter(models.GeneralExpense.id == eid).first()
    if not e:
        raise HTTPException(404, "Расход не найден")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(e, k, v)
    db.commit()
    db.refresh(e)
    return e


@app.delete("/general-expenses/{eid}")
async def delete_general_expense(eid: int, db: Session = Depends(database.get_db)):
    e = db.query(models.GeneralExpense).filter(models.GeneralExpense.id == eid).first()
    if not e:
        raise HTTPException(404, "Расход не найден")
    db.delete(e)
    db.commit()
    return {"ok": True}


# =====================
# ФИНАНСЫ
# =====================

@app.get("/financial/summary")
async def financial_summary(db: Session = Depends(database.get_db)):
    total_debit = db.query(func.sum(models.Payment.amount)).filter(
        models.Payment.status == "paid"
    ).scalar() or 0
    total_contractor = db.query(func.sum(models.ContractorWork.total_price)).filter(
        models.ContractorWork.status.in_(["completed", "paid"])
    ).scalar() or 0
    total_material = db.query(func.sum(models.MaterialPurchase.total_price)).filter(
        models.MaterialPurchase.status == "delivered"
    ).scalar() or 0
    total_general = db.query(func.sum(models.GeneralExpense.amount)).scalar() or 0
    total_credit = total_contractor + total_material + total_general

    objects = db.query(models.Object).options(
        joinedload(models.Object.estimates).joinedload(models.Estimate.items),
        joinedload(models.Object.payments),
        joinedload(models.Object.contractor_works),
        joinedload(models.Object.material_purchases),
    ).all()

    by_objects = []
    for obj in objects:
        fin = calc_object_financials_from_loaded(obj)
        if fin["total_debit"] > 0 or fin["total_credit"] > 0:
            by_objects.append({
                "object_id": obj.id, "object_name": obj.name,
                "client_name": obj.client_name, **fin,
            })

    by_month = []
    for i in range(6):
        d = date.today() - timedelta(days=30 * i)
        ms = d.strftime("%Y-%m")
        md = db.query(func.sum(models.Payment.amount)).filter(
            models.Payment.status == "paid", models.Payment.date.like(f"{ms}%"),
        ).scalar() or 0
        mc_query = db.query(func.sum(models.ContractorWork.total_price)).filter(
            models.ContractorWork.status.in_(["completed", "paid"]),
        )
        if hasattr(models.ContractorWork, 'created_at'):
            mc_query = mc_query.filter(models.ContractorWork.created_at.like(f"{ms}%"))
        mc = mc_query.scalar() or 0
        mm = db.query(func.sum(models.MaterialPurchase.total_price)).filter(
            models.MaterialPurchase.status == "delivered",
            models.MaterialPurchase.date.like(f"{ms}%"),
        ).scalar() or 0
        by_month.append({
            "month": ms, "debit": round(float(md), 2),
            "contractor": round(float(mc), 2), "material": round(float(mm), 2),
            "credit": round(float(mc + mm), 2), "profit": round(float(md - mc - mm), 2),
        })

    cm = date.today().strftime("%Y-%m")
    plan = db.query(models.FinancialPlan).filter(models.FinancialPlan.month == cm).first()

    contractor_debts_raw = (
        db.query(
            models.Contractor.id, models.Contractor.name,
            func.sum(models.ContractorWork.total_price - models.ContractorWork.advance).label("debt"),
        )
        .join(models.ContractorWork, models.ContractorWork.contractor_id == models.Contractor.id)
        .filter(models.ContractorWork.status.in_(["completed", "in_progress"]))
        .group_by(models.Contractor.id, models.Contractor.name)
        .having(func.sum(models.ContractorWork.total_price - models.ContractorWork.advance) > 0)
        .all()
    )
    contractor_debts = [
        {"contractor_id": r.id, "name": r.name, "debt": round(float(r.debt), 2)}
        for r in contractor_debts_raw
    ]

    client_debts = []
    for obj in objects:
        paid = sum(p.amount for p in obj.payments if p.status == "paid")
        estimate_total = sum(
            i.total_price for est in obj.estimates for i in est.items
        )
        debt = estimate_total - paid
        if debt > 0:
            client_debts.append({
                "object_id": obj.id, "object_name": obj.name,
                "client_name": obj.client_name, "debt": round(float(debt), 2),
            })

    return {
        "total_debit": round(float(total_debit), 2),
        "total_contractor": round(float(total_contractor), 2),
        "total_material": round(float(total_material), 2),
        "total_general": round(float(total_general), 2),
        "total_credit": round(float(total_credit), 2),
        "total_profit": round(float(total_debit - total_credit), 2),
        "by_objects": by_objects,
        "by_month": sorted(by_month, key=lambda x: x["month"]),
        "planned_income": plan.planned_income if plan else 0,
        "planned_expenses": plan.planned_expenses if plan else 0,
        "planned_profit": (plan.planned_income - plan.planned_expenses) if plan else 0,
        "contractor_debts": contractor_debts,
        "client_debts": client_debts,
    }


@app.post("/financial/plans/")
async def save_plan(data: schemas.FinancialPlanCreate, db: Session = Depends(database.get_db)):
    existing = db.query(models.FinancialPlan).filter(models.FinancialPlan.month == data.month).first()
    if existing:
        existing.planned_income = data.planned_income
        existing.planned_expenses = data.planned_expenses
        existing.notes = data.notes
        db.commit()
        db.refresh(existing)
        return existing
    plan = models.FinancialPlan(**data.model_dump())
    db.add(plan)
    db.commit()
    db.refresh(plan)
    return plan


# =====================
# ДАШБОРД
# =====================

@app.get("/dashboard/")
async def dashboard(db: Session = Depends(database.get_db)):
    current_date = today_str()

    def _task_query():
        return db.query(models.Task).options(joinedload(models.Task.object))

    overdue = _task_query().filter(
        models.Task.deadline < current_date, models.Task.status != "Done"
    ).all()
    today_tasks = _task_query().filter(
        models.Task.deadline == current_date, models.Task.status != "Done"
    ).all()
    week_end = (date.today() + timedelta(days=7)).isoformat()
    week_tasks = _task_query().filter(
        models.Task.deadline >= current_date,
        models.Task.deadline <= week_end,
        models.Task.status != "Done",
    ).order_by(models.Task.deadline).limit(10).all()

    status_counts = db.query(
        models.Object.status, func.count(models.Object.id)
    ).group_by(models.Object.status).all()

    pending = db.query(func.sum(models.Payment.amount)).filter(
        models.Payment.status == "pending"
    ).scalar() or 0
    paid = db.query(func.sum(models.Payment.amount)).filter(
        models.Payment.status == "paid"
    ).scalar() or 0

    active = db.query(models.Object).filter(models.Object.status != "Готово").limit(5).all()

    def task_to_dict(t):
        return {
            "id": t.id, "title": t.title, "deadline": t.deadline,
            "object_name": t.object.name if t.object else "",
        }

    return {
        "today_tasks": [task_to_dict(t) for t in today_tasks],
        "overdue_tasks": [task_to_dict(t) for t in overdue],
        "week_tasks": [task_to_dict(t) for t in week_tasks],
        "status_counts": {s: c for s, c in status_counts},
        "finance": {"pending": round(float(pending), 2), "paid": round(float(paid), 2)},
        "active_objects": [
            {"id": o.id, "name": o.name, "client_name": o.client_name, "status": o.status}
            for o in active
        ],
    }


# =====================
# ГЛАВНАЯ
# =====================

@app.get("/")
async def root():
    idx = FRONTEND_DIR / "index.html"
    if idx.exists():
        return FileResponse(str(idx))
    return {"message": "CRM v3.2", "docs": "/docs"}


# =====================
# СМЕТЫ ПОДРЯДЧИКА
# =====================

@app.post("/contractor-estimates/")
async def create_contractor_estimate(data: schemas.ContractorEstimateCreate, db: Session = Depends(database.get_db)):
    obj = db.query(models.Object).filter(models.Object.id == data.object_id).first()
    if not obj:
        raise HTTPException(404, "Объект не найден")
    contractor = db.query(models.Contractor).filter(models.Contractor.id == data.contractor_id).first()
    if not contractor:
        raise HTTPException(404, "Подрядчик не найден")
    name = data.name or f"Смета работ: {obj.name}"
    estimate = models.ContractorEstimate(
        object_id=data.object_id, contractor_id=data.contractor_id,
        name=name, status=data.status, notes=data.notes, created_at=now_str(),
    )
    db.add(estimate)
    db.flush()
    total = 0
    for item_data in data.items:
        item_total = round(item_data.quantity * item_data.price_per_unit, 2)
        db.add(models.ContractorEstimateItem(
            estimate_id=estimate.id, name=item_data.name, unit=item_data.unit,
            quantity=item_data.quantity, price_per_unit=item_data.price_per_unit,
            total_price=item_total,
        ))
        total += item_total
    estimate.total_sum = round(total, 2)
    db.commit()
    db.refresh(estimate)
    return build_contractor_estimate_response(estimate)


@app.get("/contractor-estimates/")
async def list_contractor_estimates(
    object_id: Optional[int] = None,
    contractor_id: Optional[int] = None,
    status: Optional[str] = None,
    db: Session = Depends(database.get_db),
):
    q = db.query(models.ContractorEstimate).options(
        joinedload(models.ContractorEstimate.contractor),
        joinedload(models.ContractorEstimate.items),
    )
    if object_id:
        q = q.filter(models.ContractorEstimate.object_id == object_id)
    if contractor_id:
        q = q.filter(models.ContractorEstimate.contractor_id == contractor_id)
    if status:
        q = q.filter(models.ContractorEstimate.status == status)
    return [build_contractor_estimate_response(e) for e in q.order_by(models.ContractorEstimate.id.desc()).all()]


@app.get("/contractor-estimates/{eid}")
async def get_contractor_estimate(eid: int, db: Session = Depends(database.get_db)):
    e = db.query(models.ContractorEstimate).options(
        joinedload(models.ContractorEstimate.contractor),
        joinedload(models.ContractorEstimate.items),
    ).filter(models.ContractorEstimate.id == eid).first()
    if not e:
        raise HTTPException(404, "Смета не найдена")
    return build_contractor_estimate_response(e)


@app.put("/contractor-estimates/{eid}")
async def update_contractor_estimate(eid: int, data: schemas.ContractorEstimateUpdate, db: Session = Depends(database.get_db)):
    e = db.query(models.ContractorEstimate).filter(models.ContractorEstimate.id == eid).first()
    if not e:
        raise HTTPException(404, "Смета не найдена")
    update_data = data.model_dump(exclude_unset=True)
    items_data = update_data.pop("items", None)
    for k, v in update_data.items():
        setattr(e, k, v)
    if items_data is not None:
        db.query(models.ContractorEstimateItem).filter(
            models.ContractorEstimateItem.estimate_id == eid
        ).delete()
        total = 0
        for item_raw in items_data:
            if isinstance(item_raw, dict):
                i_name, i_unit = item_raw["name"], item_raw["unit"]
                i_qty, i_ppu = item_raw["quantity"], item_raw["price_per_unit"]
            else:
                i_name, i_unit = item_raw.name, item_raw.unit
                i_qty, i_ppu = item_raw.quantity, item_raw.price_per_unit
            i_total = round(i_qty * i_ppu, 2)
            db.add(models.ContractorEstimateItem(
                estimate_id=eid, name=i_name, unit=i_unit,
                quantity=i_qty, price_per_unit=i_ppu, total_price=i_total,
            ))
            total += i_total
        e.total_sum = round(total, 2)
    db.commit()
    db.refresh(e)
    return build_contractor_estimate_response(e)


@app.delete("/contractor-estimates/{eid}")
async def delete_contractor_estimate(eid: int, db: Session = Depends(database.get_db)):
    e = db.query(models.ContractorEstimate).filter(models.ContractorEstimate.id == eid).first()
    if not e:
        raise HTTPException(404, "Смета не найдена")
    db.delete(e)
    db.commit()
    return {"ok": True}


@app.get("/contractor-estimates/{eid}/excel")
async def download_estimate_excel(eid: int, db: Session = Depends(database.get_db)):
    e = db.query(models.ContractorEstimate).options(
        joinedload(models.ContractorEstimate.contractor),
        joinedload(models.ContractorEstimate.items),
    ).filter(models.ContractorEstimate.id == eid).first()
    if not e:
        raise HTTPException(404, "Смета не найдена")
    contractor = e.contractor
    wb = Workbook()
    ws = wb.active
    ws.title = "Смета работ"
    header_font = Font(bold=True, size=14)
    subheader_font = Font(bold=True, size=12)
    bold_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    money_format = '#,##0.00" ₽"'
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    ws.merge_cells("A1:F1")
    ws["A1"] = e.name
    ws["A1"].font = header_font
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:F2")
    ws["A2"] = f"Подрядчик: {contractor.name if contractor else '—'} | Телефон: {contractor.phone if contractor else '—'}"
    ws["A2"].font = subheader_font
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A3:F3")
    ws["A3"] = f"Статус: {e.status} | Дата: {e.created_at or ''}"
    ws["A3"].alignment = Alignment(horizontal="center")
    col_headers = ["№", "Наименование работ", "Ед. изм.", "Кол-во", "Цена за ед.", "Сумма"]
    for col_idx, h in enumerate(col_headers, 1):
        cell = ws.cell(row=5, column=col_idx, value=h)
        cell.font = bold_font
        cell.alignment = center
        cell.border = thin_border
        cell.fill = header_fill
    row = 6
    for idx, item in enumerate(e.items, 1):
        ws.cell(row=row, column=1, value=idx).alignment = center
        ws.cell(row=row, column=2, value=item.name).alignment = left_wrap
        ws.cell(row=row, column=3, value=item.unit).alignment = center
        ws.cell(row=row, column=4, value=item.quantity).alignment = center
        c = ws.cell(row=row, column=5, value=item.price_per_unit)
        c.number_format = money_format
        c.alignment = center
        c = ws.cell(row=row, column=6, value=item.total_price)
        c.number_format = money_format
        c.alignment = center
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = thin_border
        row += 1
    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    ws.cell(row=row, column=1, value="ИТОГО:").font = bold_font
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="right")
    c = ws.cell(row=row, column=6, value=e.total_sum)
    c.font = bold_font
    c.number_format = money_format
    c.alignment = center
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"estimate_{e.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# =====================
# ПОИСК ПОЗИЦИЙ СМЕТЫ
# =====================

@app.get("/estimate-items/search")
async def search_all_estimate_items(q: str = "", db: Session = Depends(database.get_db)):
    safe_q = escape_like(q)
    items = db.query(models.EstimateItem).filter(
        models.EstimateItem.name.ilike(f"%{safe_q}%")
    ).limit(20).all()
    return [{
        "id": i.id, "estimate_id": i.estimate_id, "name": i.name,
        "unit": i.unit, "quantity": i.quantity,
        "price_per_unit": i.price_per_unit, "total_price": i.total_price,
        "contractor_price": round(i.total_price * CONTRACTOR_MARGIN, 2),
        "contractor_price_per_unit": round(i.price_per_unit * CONTRACTOR_MARGIN, 2),
    } for i in items]


# =====================
# SUPPLIERS
# =====================

@app.get("/suppliers/")
def list_suppliers(db: Session = Depends(database.get_db)):
    suppliers_list = db.query(Supplier).order_by(Supplier.id.desc()).all()
    result = []
    for s in suppliers_list:
        files = []
        for f in s.price_files:
            files.append({
                "id": f.id,
                "file_name": f.file_name,
                "file_path": f.file_path,
                "description": f.description,
                "uploaded_at": f.uploaded_at,
                "items_count": len(f.items),
                "items": [{"id": it.id, "name": it.name, "unit": it.unit,
                           "price_per_unit": it.price_per_unit, "row_number": it.row_number}
                          for it in f.items]
            })
        result.append({
            "id": s.id, "name": s.name, "phone": s.phone, "email": s.email,
            "address": s.address, "notes": s.notes, "created_at": s.created_at,
            "price_files": files,
            "total_items": sum(len(f.items) for f in s.price_files)
        })
    return result


@app.post("/suppliers/")
def create_supplier(data: dict, db: Session = Depends(database.get_db)):
    s = Supplier(
        name=data["name"], phone=data.get("phone", ""),
        email=data.get("email", ""), address=data.get("address", ""),
        notes=data.get("notes", "")
    )
    db.add(s)
    db.commit()
    db.refresh(s)
    return {"id": s.id, "name": s.name, "phone": s.phone, "email": s.email,
            "address": s.address, "notes": s.notes, "created_at": s.created_at}


@app.put("/suppliers/{supplier_id}")
def update_supplier(supplier_id: int, data: dict, db: Session = Depends(database.get_db)):
    s = db.query(Supplier).filter(Supplier.id == supplier_id).first()
    if not s:
        raise HTTPException(404, "Поставщик не найден")
    for field in ["name", "phone", "email", "address", "notes"]:
        if field in data:
            setattr(s, field, data[field])
    db.commit()
    return {"ok": True}


@app.delete("/suppliers/{supplier_id}")
def delete_supplier(supplier_id: int, db: Session = Depends(database.get_db)):
    s = db.query(Supplier).filter(Supplier.id == supplier_id).first()
    if not s:
        raise HTTPException(404, "Поставщик не найден")
    for f in s.price_files:
        if f.file_path and os.path.exists(f.file_path):
            os.remove(f.file_path)
    db.delete(s)
    db.commit()
    return Response(status_code=204)


@app.post("/suppliers/{supplier_id}/price-file/upload")
async def upload_supplier_price_file(
    supplier_id: int,
    file: UploadFile = File(...),
    description: str = "",
    db: Session = Depends(database.get_db)
):
    s = db.query(Supplier).filter(Supplier.id == supplier_id).first()
    if not s:
        raise HTTPException(404, "Поставщик не найден")

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in [".xlsx", ".xls"]:
        raise HTTPException(400, "Только Excel файлы (.xlsx, .xls)")

    safe_name = f"supplier_{supplier_id}_{int(time.time())}_{file.filename}"
    file_path = os.path.join(UPLOAD_DIR, safe_name)
    content = await file.read()
    with open(file_path, "wb") as f:
        f.write(content)

    items = parse_supplier_price_excel(content, file.filename)

    pf = SupplierPriceFile(
        supplier_id=supplier_id,
        file_name=file.filename,
        file_path=file_path,
        description=description
    )
    db.add(pf)
    db.flush()

    for item in items:
        pi = SupplierPriceItem(
            price_file_id=pf.id,
            name=item["name"],
            unit=item.get("unit", "шт"),
            price_per_unit=item.get("price_per_unit", 0),
            row_number=item.get("row_number")
        )
        db.add(pi)

    db.commit()
    return {"message": f"Загружено {len(items)} позиций", "file_id": pf.id, "items_count": len(items)}


@app.delete("/suppliers/price-files/{file_id}")
def delete_supplier_price_file(file_id: int, db: Session = Depends(database.get_db)):
    pf = db.query(SupplierPriceFile).filter(SupplierPriceFile.id == file_id).first()
    if not pf:
        raise HTTPException(404, "Файл не найден")
    if pf.file_path and os.path.exists(pf.file_path):
        os.remove(pf.file_path)
    db.delete(pf)
    db.commit()
    return Response(status_code=204)


@app.get("/suppliers/{supplier_id}/price-items/search")
def search_supplier_price_items(supplier_id: int, q: str = "", db: Session = Depends(database.get_db)):
    if len(q) < 2:
        return []
    items = db.query(SupplierPriceItem).join(SupplierPriceFile).filter(
        SupplierPriceFile.supplier_id == supplier_id,
        SupplierPriceItem.name.ilike(f"%{escape_like(q)}%")
    ).limit(20).all()
    return [{"id": i.id, "name": i.name, "unit": i.unit, "price_per_unit": i.price_per_unit,
             "row_number": i.row_number, "file_id": i.price_file_id} for i in items]


def parse_supplier_price_excel(content: bytes, filename: str) -> list:
    """Парсит Excel файл расценок поставщика"""
    import io
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(400, "Не удалось прочитать Excel файл")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_idx = None
    col_map = {}
    name_variants = ["наименование", "название", "товар", "материал", "позиция", "описание", "name"]
    unit_variants = ["ед", "ед.изм", "ед. изм", "единица", "unit", "изм"]
    price_variants = ["цена", "стоимость", "price", "руб", "цена за ед"]

    for i, row in enumerate(rows):
        vals = [str(c).lower().strip() if c else "" for c in row]
        for j, v in enumerate(vals):
            if not v:
                continue
            for nv in name_variants:
                if nv in v and "name" not in col_map:
                    col_map["name"] = j
                    header_idx = i
            for uv in unit_variants:
                if uv in v and "unit" not in col_map:
                    col_map["unit"] = j
            for pv in price_variants:
                if pv in v and "price" not in col_map:
                    col_map["price"] = j
        if "name" in col_map:
            break

    if "name" not in col_map:
        col_map = {"name": 0, "unit": 1, "price": 2}
        header_idx = -1

    items = []
    for i, row in enumerate(rows):
        if i <= (header_idx if header_idx is not None else -1):
            continue
        name_val = row[col_map["name"]] if col_map["name"] < len(row) else None
        if not name_val or not str(name_val).strip():
            continue
        name_str = str(name_val).strip()
        if any(x in name_str.lower() for x in ["итого", "всего", "total", "№ п/п"]):
            continue

        unit_val = ""
        if "unit" in col_map and col_map["unit"] < len(row):
            unit_val = str(row[col_map["unit"]] or "").strip()

        price_val = 0
        if "price" in col_map and col_map["price"] < len(row):
            try:
                pv = row[col_map["price"]]
                price_val = float(str(pv).replace(",", ".").replace(" ", "").replace("\xa0", "")) if pv else 0
            except (ValueError, TypeError):
                price_val = 0

        items.append({
            "name": name_str,
            "unit": unit_val or "шт",
            "price_per_unit": round(price_val, 2),
            "row_number": i + 1
        })

    return items


# =====================
# MATERIAL REQUESTS (ЗАЯВКИ НА МАТЕРИАЛЫ)
# =====================

@app.get("/material-requests/")
def list_material_requests(object_id: int = None, db: Session = Depends(database.get_db)):
    q = db.query(MaterialRequest).order_by(MaterialRequest.id.desc())
    if object_id:
        q = q.filter(MaterialRequest.object_id == object_id)
    reqs = q.all()
    result = []
    for r in reqs:
        obj = db.query(models.Object).filter(models.Object.id == r.object_id).first() if r.object_id else None
        sup = db.query(Supplier).filter(Supplier.id == r.supplier_id).first() if r.supplier_id else None
        result.append({
            "id": r.id, "name": r.name, "status": r.status,
            "total_sum": r.total_sum, "notes": r.notes, "created_at": r.created_at,
            "file_path": r.file_path,
            "object_id": r.object_id, "object_name": obj.name if obj else "",
            "supplier_id": r.supplier_id, "supplier_name": sup.name if sup else "",
            "items": [{"id": it.id, "name": it.name, "unit": it.unit, "quantity": it.quantity,
                       "price_per_unit": it.price_per_unit, "total_price": it.total_price,
                       "supplier_price_item_id": it.supplier_price_item_id} for it in r.items]
        })
    return result


@app.post("/material-requests/")
def create_material_request(data: dict, db: Session = Depends(database.get_db)):
    mr = MaterialRequest(
        object_id=data.get("object_id"),
        supplier_id=data.get("supplier_id"),
        name=data.get("name", "Заявка на материалы"),
        status=data.get("status", "Черновик"),
        notes=data.get("notes", "")
    )
    db.add(mr)
    db.flush()

    total = 0
    for item in data.get("items", []):
        name = item.get("name", "").strip()
        if not name:
            continue
        qty = float(item.get("quantity", 1) or 1)
        price = float(item.get("price_per_unit", 0) or 0)
        tp = round(qty * price, 2)
        total += tp
        mi = MaterialRequestItem(
            request_id=mr.id, name=name,
            unit=item.get("unit", "шт"),
            quantity=qty, price_per_unit=price, total_price=tp,
            supplier_price_item_id=item.get("supplier_price_item_id")
        )
        db.add(mi)

    mr.total_sum = round(total, 2)
    db.commit()
    db.refresh(mr)
    return {"id": mr.id, "total_sum": mr.total_sum, "message": f"Создана заявка с {len(mr.items)} позициями"}


@app.put("/material-requests/{req_id}")
def update_material_request(req_id: int, data: dict, db: Session = Depends(database.get_db)):
    mr = db.query(MaterialRequest).filter(MaterialRequest.id == req_id).first()
    if not mr:
        raise HTTPException(404, "Заявка не найдена")

    for field in ["name", "status", "notes", "object_id", "supplier_id"]:
        if field in data:
            setattr(mr, field, data[field])

    if "items" in data:
        db.query(MaterialRequestItem).filter(MaterialRequestItem.request_id == req_id).delete()
        total = 0
        for item in data["items"]:
            name = item.get("name", "").strip()
            if not name:
                continue
            qty = float(item.get("quantity", 1) or 1)
            price = float(item.get("price_per_unit", 0) or 0)
            tp = round(qty * price, 2)
            total += tp
            mi = MaterialRequestItem(
                request_id=mr.id, name=name,
                unit=item.get("unit", "шт"),
                quantity=qty, price_per_unit=price, total_price=tp,
                supplier_price_item_id=item.get("supplier_price_item_id")
            )
            db.add(mi)
        mr.total_sum = round(total, 2)

    db.commit()
    return {"ok": True}


@app.delete("/material-requests/{req_id}")
def delete_material_request(req_id: int, db: Session = Depends(database.get_db)):
    mr = db.query(MaterialRequest).filter(MaterialRequest.id == req_id).first()
    if not mr:
        raise HTTPException(404, "Заявка не найдена")
    if mr.file_path and os.path.exists(mr.file_path):
        os.remove(mr.file_path)
    db.delete(mr)
    db.commit()
    return Response(status_code=204)


@app.get("/material-requests/{req_id}/export")
def export_material_request(req_id: int, db: Session = Depends(database.get_db)):
    mr = db.query(MaterialRequest).filter(MaterialRequest.id == req_id).first()
    if not mr:
        raise HTTPException(404, "Заявка не найдена")

    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заявка"

    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Заголовок
    ws.merge_cells('A1:F1')
    ws['A1'] = mr.name
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    obj = db.query(models.Object).filter(models.Object.id == mr.object_id).first() if mr.object_id else None
    sup = db.query(Supplier).filter(Supplier.id == mr.supplier_id).first() if mr.supplier_id else None

    row = 2
    if obj:
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = f"Объект: {obj.name}"
        row += 1
    if sup:
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = f"Поставщик: {sup.name}" + (f" | Тел: {sup.phone}" if sup.phone else "")
        row += 1

    row += 1
    headers = ["№", "Наименование", "Ед. изм.", "Кол-во", "Цена за ед.", "Сумма"]
    widths = [5, 45, 10, 10, 15, 15]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=i+1, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[chr(65 + i)].width = widths[i]

    for idx, item in enumerate(mr.items, 1):
        row += 1
        vals = [idx, item.name, item.unit, item.quantity, item.price_per_unit, item.total_price]
        for i, v in enumerate(vals):
            cell = ws.cell(row=row, column=i+1, value=v)
            cell.border = border
            if i >= 3:
                cell.number_format = '#,##0.00'

    # Итого
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "ИТОГО:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'A{row}'].alignment = Alignment(horizontal='right')
    ws[f'A{row}'].border = border
    total_cell = ws.cell(row=row, column=6, value=mr.total_sum)
    total_cell.font = Font(bold=True)
    total_cell.border = border
    total_cell.number_format = '#,##0.00'

    file_name = f"material_request_{mr.id}.xlsx"
    file_path = os.path.join(UPLOAD_DIR, file_name)
    wb.save(file_path)
    mr.file_path = file_path
    db.commit()

    return FileResponse(file_path, filename=f"{mr.name}.xlsx",
                       media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
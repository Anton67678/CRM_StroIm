import os
import uuid
import time
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import List, Optional

from fastapi import FastAPI, Depends, HTTPException, Query, UploadFile, File, status, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO

import models
import schemas
import database

# ===== Инициализация =====
models.Base.metadata.create_all(bind=database.engine)

app = FastAPI(title="CRM Repair System", version="4.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], allow_credentials=True,
    allow_methods=["*"], allow_headers=["*"],
)

BASE_DIR = Path(__file__).parent
FRONTEND_DIR = BASE_DIR / "frontend"
UPLOADS_DIR = BASE_DIR / "uploads"
UPLOADS_DIR.mkdir(exist_ok=True)
UPLOAD_DIR = str(UPLOADS_DIR)

CONTRACTOR_MARGIN = float(os.getenv("CONTRACTOR_MARGIN", "0.5"))
MAX_FILE_SIZE = int(os.getenv("MAX_FILE_SIZE", str(50 * 1024 * 1024)))

if FRONTEND_DIR.exists():
    app.mount("/static", StaticFiles(directory=str(FRONTEND_DIR)), name="static")


# ===== Утилиты =====
def now_str(): return datetime.now().strftime("%Y-%m-%d %H:%M:%S")
def today_str(): return date.today().isoformat()
def escape_like(s: str) -> str: return s.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
def safe_float(val, default=0.0):
    try: return float(val or default)
    except (ValueError, TypeError): return default
def round2(n): return round(float(n or 0), 2)


# ===== Финансы объекта =====
def calc_object_financials_from_loaded(obj: models.Object) -> dict:
    # Дебет: оплаченные платежи + акты со статусом "Оплачен"
    debit_payments = sum(p.amount for p in obj.payments if p.status == "paid")
    debit_acts = sum(a.total_sum for a in obj.acts if a.status == "Оплачен")
    debit = debit_payments + debit_acts

    # Расходы: сметы подрядчика (paid) + доп. работы (paid) + старые работы (completed/paid)
    contractor_est = sum(
        est.total_sum for est in obj.contractor_estimates if est.status == "paid"
    )
    extra_work = sum(w.total_price for w in obj.extra_works if w.status == "paid")
    old_work = sum(w.total_price for w in obj.contractor_works if w.status in ("completed", "paid"))
    contractor_total = contractor_est + extra_work + old_work

    # Материалы: заявки (Оплачено) + закупки (delivered)
    mat_requests = sum(
        mr.total_sum for mr in obj.material_requests if mr.status == "Оплачено"
    )
    mat_purchases = sum(m.total_price for m in obj.material_purchases if m.status == "delivered")
    material_total = mat_requests + mat_purchases

    # Акты: сумма подписанных актов
    acts_total = debit_acts

    credit = contractor_total + material_total
    profit = debit - credit
    margin = (profit / debit * 100) if debit > 0 else 0
    return {
        "total_debit": round2(debit),
        "contractor_expenses": round2(contractor_total),
        "material_expenses": round2(material_total),
        "acts_total": round2(acts_total),
        "total_credit": round2(credit),
        "profit": round2(profit),
        "margin_percent": round(margin, 1),
    }


def _object_joinedload_options():
    return [
        joinedload(models.Object.estimates).joinedload(models.Estimate.items),
        joinedload(models.Object.payments),
        joinedload(models.Object.contractor_works).joinedload(models.ContractorWork.contractor),
        joinedload(models.Object.contractor_works).joinedload(models.ContractorWork.tools),
        joinedload(models.Object.material_purchases).joinedload(models.MaterialPurchase.material),
        joinedload(models.Object.documents),
        joinedload(models.Object.communications),
        joinedload(models.Object.tasks),
        joinedload(models.Object.contractor_estimates).joinedload(models.ContractorEstimate.items).joinedload(models.ContractorEstimateItem.estimate_item),
        joinedload(models.Object.contractor_estimates).joinedload(models.ContractorEstimate.contractor),
        joinedload(models.Object.extra_works),
        joinedload(models.Object.material_requests).joinedload(models.MaterialRequest.items),
        joinedload(models.Object.acts),
    ]


def parse_estimate_excel(file_path: str):
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        items = []
        header_row = None
        headers = {}
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=False), 1):
            row_values = [str(c.value).strip().lower() if c.value else "" for c in row]
            for col_idx, val in enumerate(row_values):
                if "наименован" in val or "работ" in val:
                    headers["name"] = col_idx; header_row = row_idx
                elif "ед" in val and ("изм" in val or "." in val):
                    headers["unit"] = col_idx
                elif "объём" in val or "объем" in val or "кол" in val:
                    headers["quantity"] = col_idx
                elif "цена" in val and "ед" in val:
                    headers["price"] = col_idx
                elif "сумм" in val or "стоим" in val:
                    headers["total"] = col_idx
            if "name" in headers: break
        if not header_row:
            headers = {"name": 0, "unit": 1, "quantity": 2, "price": 3, "total": 4}
            header_row = 1
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            row = list(row) + [None] * 10
            name_val = row[headers.get("name", 0)]
            if not name_val or str(name_val).strip() == "": continue
            unit = str(row[headers.get("unit", 1)] or "").strip()
            quantity = safe_float(row[headers.get("quantity", 2)])
            price = safe_float(row[headers.get("price", 3)])
            total = safe_float(row[headers.get("total", 4)])
            if total == 0 and quantity > 0 and price > 0: total = quantity * price
            items.append({"name": str(name_val).strip(), "unit": unit, "quantity": quantity,
                          "price_per_unit": price, "total_price": total})
        return items
    except Exception as e:
        raise HTTPException(400, f"Ошибка парсинга Excel: {str(e)}")


def parse_supplier_price_excel(content: bytes, filename: str) -> list:
    import io, openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(400, "Не удалось прочитать Excel файл")
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return []
    header_idx = None; col_map = {}
    name_variants = ["наименование", "название", "товар", "материал", "позиция", "описание", "name"]
    unit_variants = ["ед", "ед.изм", "ед. изм", "единица", "unit", "изм"]
    price_variants = ["цена", "стоимость", "price", "руб", "цена за ед"]
    for i, row in enumerate(rows):
        vals = [str(c).lower().strip() if c else "" for c in row]
        for j, v in enumerate(vals):
            if not v: continue
            for nv in name_variants:
                if nv in v and "name" not in col_map: col_map["name"] = j; header_idx = i
            for uv in unit_variants:
                if uv in v and "unit" not in col_map: col_map["unit"] = j
            for pv in price_variants:
                if pv in v and "price" not in col_map: col_map["price"] = j
        if "name" in col_map: break
    if "name" not in col_map: col_map = {"name": 0, "unit": 1, "price": 2}; header_idx = -1
    items = []
    for i, row in enumerate(rows):
        if i <= (header_idx if header_idx is not None else -1): continue
        name_val = row[col_map["name"]] if col_map["name"] < len(row) else None
        if not name_val or not str(name_val).strip(): continue
        name_str = str(name_val).strip()
        if any(x in name_str.lower() for x in ["итого", "всего", "total", "№ п/п"]): continue
        unit_val = ""
        if "unit" in col_map and col_map["unit"] < len(row):
            unit_val = str(row[col_map["unit"]] or "").strip()
        price_val = 0
        if "price" in col_map and col_map["price"] < len(row):
            try:
                pv = row[col_map["price"]]
                price_val = float(str(pv).replace(",", ".").replace(" ", "").replace("\xa0", "")) if pv else 0
            except: price_val = 0
        items.append({"name": name_str, "unit": unit_val or "шт", "price_per_unit": round(price_val, 2), "row_number": i + 1})
    return items


# =====================
# RESPONSE BUILDERS
# =====================

def build_object_response(obj: models.Object) -> dict:
    fin = calc_object_financials_from_loaded(obj)
    return {
        "id": obj.id, "name": obj.name, "client_name": obj.client_name,
        "client_phone": obj.client_phone, "client_email": obj.client_email,
        "client_address": obj.client_address, "status": obj.status,
        "created_at": obj.created_at, "notes": obj.notes,
        "estimates": [
            {"id": e.id, "object_id": e.object_id, "name": e.name,
             "file_path": e.file_path, "created_at": e.created_at,
             "items": [{"id": i.id, "estimate_id": i.estimate_id, "name": i.name,
                        "unit": i.unit, "quantity": i.quantity,
                        "price_per_unit": i.price_per_unit, "total_price": i.total_price}
                       for i in e.items]}
            for e in obj.estimates
        ],
        "payments": [
            {"id": p.id, "object_id": p.object_id, "amount": p.amount,
             "status": p.status, "description": p.description, "date": p.date}
            for p in obj.payments
        ],
        "contractor_works": [build_work_response(w) for w in obj.contractor_works],
        "material_purchases": [
            {"id": m.id, "object_id": m.object_id, "material_id": m.material_id,
             "quantity": m.quantity, "total_price": m.total_price,
             "supplier": m.supplier, "date": m.date, "status": m.status, "notes": m.notes,
             "material": {"id": m.material.id, "name": m.material.name,
                          "unit": m.material.unit, "price_per_unit": m.material.price_per_unit,
                          "description": m.material.description} if m.material else None}
            for m in obj.material_purchases
        ],
        "documents": [
            {"id": d.id, "object_id": d.object_id, "doc_type": d.doc_type,
             "name": d.name, "file_path": d.file_path, "created_at": d.created_at}
            for d in obj.documents
        ],
        "communications": [
            {"id": c.id, "object_id": c.object_id, "type": c.type,
             "description": c.description, "date": c.date}
            for c in obj.communications
        ],
        "tasks": [
            {"id": t.id, "object_id": t.object_id, "title": t.title,
             "description": t.description, "status": t.status,
             "deadline": t.deadline, "created_at": t.created_at}
            for t in obj.tasks
        ],
        "contractor_estimates": [
            build_contractor_estimate_response(e) for e in obj.contractor_estimates
        ],
        "financials": {"object_id": obj.id, "object_name": obj.name,
                       "client_name": obj.client_name, **fin},
    }


def build_work_response(w: models.ContractorWork) -> dict:
    return {
        "id": w.id, "object_id": w.object_id, "contractor_id": w.contractor_id,
        "estimate_item_id": w.estimate_item_id, "description": w.description,
        "unit": w.unit, "quantity": w.quantity, "price_per_unit": w.price_per_unit,
        "total_price": w.total_price, "advance": w.advance, "deadline": w.deadline,
        "status": w.status, "notes": w.notes, "created_at": w.created_at,
        "contractor": {"id": w.contractor.id, "name": w.contractor.name,
                       "phone": w.contractor.phone, "specialization": w.contractor.specialization,
                       "notes": w.contractor.notes} if w.contractor else None,
        "tools": [{"id": t.id, "name": t.name, "serial_number": t.serial_number,
                   "purchase_price": t.purchase_price, "purchase_date": t.purchase_date,
                   "status": t.status, "contractor_id": t.contractor_id,
                   "object_id": t.object_id, "notes": t.notes} for t in w.tools],
    }


def build_contractor_estimate_response(e: models.ContractorEstimate) -> dict:
    items = []
    for i in e.items:
        client_ppu = i.estimate_item.price_per_unit if i.estimate_item else 0
        client_total = i.estimate_item.total_price if i.estimate_item else 0
        items.append({
            "id": i.id, "estimate_id": i.estimate_id,
            "estimate_item_id": i.estimate_item_id,
            "name": i.name, "unit": i.unit, "quantity": i.quantity,
            "price_per_unit": i.price_per_unit, "total_price": i.total_price,
            "client_price_per_unit": client_ppu, "client_total_price": client_total,
        })
    return {
        "id": e.id, "object_id": e.object_id, "contractor_id": e.contractor_id,
        "name": e.name, "status": e.status, "total_sum": e.total_sum,
        "created_at": e.created_at, "completed_at": e.completed_at,
        "paid_at": e.paid_at, "notes": e.notes,
        "items": items,
        "contractor": {"id": e.contractor.id, "name": e.contractor.name,
                       "phone": e.contractor.phone, "specialization": e.contractor.specialization}
        if e.contractor else None,
    }


# =====================
# ОБЪЕКТЫ
# =====================

@app.post("/objects/", response_model=schemas.ObjectFullResponse)
async def create_object(obj: schemas.ObjectCreate, db: Session = Depends(database.get_db)):
    db_obj = models.Object(**obj.model_dump(), created_at=now_str())
    db.add(db_obj); db.commit(); db.refresh(db_obj)
    return build_object_response(db_obj)


@app.get("/objects/", response_model=List[schemas.ObjectFullResponse])
async def list_objects(status: Optional[str] = None, search: Optional[str] = None,
                       db: Session = Depends(database.get_db)):
    q = db.query(models.Object).options(*_object_joinedload_options())
    if status: q = q.filter(models.Object.status == status)
    if search:
        s = escape_like(search)
        q = q.filter(models.Object.name.ilike(f"%{s}%") | models.Object.client_name.ilike(f"%{s}%"))
    return [build_object_response(o) for o in q.order_by(models.Object.id.desc()).all()]


@app.get("/objects/{obj_id}", response_model=schemas.ObjectFullResponse)
async def get_object(obj_id: int, db: Session = Depends(database.get_db)):
    obj = db.query(models.Object).options(*_object_joinedload_options()).filter(models.Object.id == obj_id).first()
    if not obj: raise HTTPException(404, "Объект не найден")
    return build_object_response(obj)


@app.put("/objects/{obj_id}", response_model=schemas.ObjectFullResponse)
async def update_object(obj_id: int, data: schemas.ObjectUpdate, db: Session = Depends(database.get_db)):
    obj = db.query(models.Object).options(*_object_joinedload_options()).filter(models.Object.id == obj_id).first()
    if not obj: raise HTTPException(404, "Объект не найден")
    for k, v in data.model_dump(exclude_unset=True).items(): setattr(obj, k, v)
    db.commit(); db.refresh(obj)
    return build_object_response(obj)


@app.delete("/objects/{obj_id}")
async def delete_object(obj_id: int, db: Session = Depends(database.get_db)):
    obj = db.query(models.Object).filter(models.Object.id == obj_id).first()
    if not obj: raise HTTPException(404, "Объект не найден")
    db.delete(obj); db.commit()
    return {"ok": True}


# =====================
# СМЕТЫ ОБЪЕКТА
# =====================

@app.post("/objects/{obj_id}/estimate/upload")
async def upload_estimate(obj_id: int, file: UploadFile = File(...), db: Session = Depends(database.get_db)):
    obj = db.query(models.Object).filter(models.Object.id == obj_id).first()
    if not obj: raise HTTPException(404, "Объект не найден")
    ext = Path(file.filename).suffix.lower()
    if ext not in (".xlsx", ".xls", ".pdf"): raise HTTPException(400, "Поддерживаются только Excel и PDF")
    content = await file.read()
    if len(content) > MAX_FILE_SIZE: raise HTTPException(413, "Файл слишком большой")
    safe_name = f"estimate_{obj_id}_{uuid.uuid4().hex[:8]}{ext}"
    file_path = UPLOADS_DIR / safe_name
    with open(file_path, "wb") as f: f.write(content)
    estimate_count = db.query(models.Estimate).filter(models.Estimate.object_id == obj_id).count()
    estimate = models.Estimate(object_id=obj_id, name=f"Смета №{estimate_count + 1}",
                               file_path=str(file_path), created_at=now_str())
    db.add(estimate); db.flush()
    items_count = 0
    if ext in (".xlsx", ".xls"):
        for item in parse_estimate_excel(str(file_path)):
            db.add(models.EstimateItem(estimate_id=estimate.id, **item))
            items_count += 1
    db.commit()
    return {"estimate_id": estimate.id, "items_count": items_count,
            "message": f"Смета загружена. Распознано позиций: {items_count}"}


@app.delete("/estimates/{estimate_id}")
async def delete_estimate(estimate_id: int, db: Session = Depends(database.get_db)):
    est = db.query(models.Estimate).filter(models.Estimate.id == estimate_id).first()
    if not est: raise HTTPException(404)
    if est.file_path and os.path.exists(est.file_path): os.remove(est.file_path)
    db.delete(est); db.commit()
    return {"ok": True}


@app.get("/objects/{obj_id}/estimate-items/search")
async def search_estimate_items(obj_id: int, q: str = "", db: Session = Depends(database.get_db)):
    items = db.query(models.EstimateItem).join(models.Estimate).filter(
        models.Estimate.object_id == obj_id,
        models.EstimateItem.name.ilike(f"%{escape_like(q)}%"),
    ).limit(20).all()
    return [{
        "id": i.id, "estimate_id": i.estimate_id, "name": i.name,
        "unit": i.unit, "quantity": i.quantity,
        "price_per_unit": i.price_per_unit, "total_price": i.total_price,
        "contractor_price": round(i.total_price * CONTRACTOR_MARGIN, 2),
        "contractor_price_per_unit": round(i.price_per_unit * CONTRACTOR_MARGIN, 2),
    } for i in items]


@app.get("/estimate-items/search")
async def search_all_estimate_items(q: str = "", db: Session = Depends(database.get_db)):
    items = db.query(models.EstimateItem).filter(
        models.EstimateItem.name.ilike(f"%{escape_like(q)}%")
    ).limit(20).all()
    return [{
        "id": i.id, "estimate_id": i.estimate_id, "name": i.name,
        "unit": i.unit, "quantity": i.quantity,
        "price_per_unit": i.price_per_unit, "total_price": i.total_price,
        "contractor_price": round(i.total_price * CONTRACTOR_MARGIN, 2),
        "contractor_price_per_unit": round(i.price_per_unit * CONTRACTOR_MARGIN, 2),
    } for i in items]


@app.get("/estimate-items/locked")
async def get_locked_estimate_items(object_id: int, db: Session = Depends(database.get_db)):
    """Возвращает ID позиций из основной сметы, которые уже используются в сметах подрядчика
    со статусами, блокирующими повторное назначение."""
    estimates = db.query(models.ContractorEstimate).options(
        joinedload(models.ContractorEstimate.contractor),
        joinedload(models.ContractorEstimate.items)
    ).filter(
        models.ContractorEstimate.object_id == object_id,
        models.ContractorEstimate.status.in_(["in_progress", "completed", "paid", "Акт"])
    ).all()
    
    locked_ids = set()
    details = []
    for est in estimates:
        for item in est.items:
            if item.estimate_item_id and item.estimate_item_id not in locked_ids:
                locked_ids.add(item.estimate_item_id)
                details.append({
                    "estimate_item_id": item.estimate_item_id,
                    "contractor_name": est.contractor.name if est.contractor else "—",
                    "estimate_name": est.name,
                    "status": est.status
                })
    return {"locked_ids": list(locked_ids), "details": details}


# =====================
# ДОКУМЕНТЫ
# =====================

@app.post("/objects/{obj_id}/document/upload")
async def upload_object_document(obj_id: int, doc_type: str = "act",
                                  file: UploadFile = File(...), db: Session = Depends(database.get_db)):
    obj = db.query(models.Object).filter(models.Object.id == obj_id).first()
    if not obj: raise HTTPException(404)
    content = await file.read()
    ext = Path(file.filename).suffix.lower()
    safe_name = f"doc_{obj_id}_{uuid.uuid4().hex[:8]}{ext}"
    file_path = UPLOADS_DIR / safe_name
    with open(file_path, "wb") as f: f.write(content)
    doc = models.ObjectDocument(object_id=obj_id, doc_type=doc_type, name=file.filename,
                                file_path=str(file_path), created_at=now_str())
    db.add(doc); db.commit(); db.refresh(doc)
    return {"id": doc.id, "name": doc.name}


@app.get("/uploads/{filename}")
async def get_uploaded_file(filename: str):
    file_path = (UPLOADS_DIR / filename).resolve()
    if not str(file_path).startswith(str(UPLOADS_DIR.resolve())): raise HTTPException(403)
    if not file_path.exists(): raise HTTPException(404)
    return FileResponse(str(file_path))


# =====================
# ПЛАТЕЖИ КЛИЕНТОВ
# =====================

@app.post("/payments/", response_model=schemas.PaymentResponse)
async def create_payment(data: schemas.PaymentCreate, db: Session = Depends(database.get_db)):
    payload = data.model_dump()
    if not payload.get("date"): payload["date"] = today_str()
    p = models.Payment(**payload); db.add(p); db.commit(); db.refresh(p)
    return p


@app.get("/payments/")
async def list_payments(object_id: Optional[int] = None, db: Session = Depends(database.get_db)):
    q = db.query(models.Payment)
    if object_id: q = q.filter(models.Payment.object_id == object_id)
    return q.order_by(models.Payment.id.desc()).all()


@app.put("/payments/{pay_id}", response_model=schemas.PaymentResponse)
async def update_payment(pay_id: int, data: schemas.PaymentUpdate, db: Session = Depends(database.get_db)):
    p = db.query(models.Payment).filter(models.Payment.id == pay_id).first()
    if not p: raise HTTPException(404)
    for k, v in data.model_dump(exclude_unset=True).items(): setattr(p, k, v)
    db.commit(); db.refresh(p)
    return p


@app.delete("/payments/{pay_id}")
async def delete_payment(pay_id: int, db: Session = Depends(database.get_db)):
    p = db.query(models.Payment).filter(models.Payment.id == pay_id).first()
    if not p: raise HTTPException(404)
    db.delete(p); db.commit()
    return {"ok": True}


# =====================
# ПОДРЯДЧИКИ
# =====================

@app.post("/contractors/", response_model=schemas.ContractorResponse)
async def create_contractor(data: schemas.ContractorCreate, db: Session = Depends(database.get_db)):
    c = models.Contractor(**data.model_dump()); db.add(c); db.commit(); db.refresh(c)
    return c


@app.get("/contractors/", response_model=List[schemas.ContractorResponse])
async def list_contractors(db: Session = Depends(database.get_db)):
    return db.query(models.Contractor).all()


@app.put("/contractors/{cid}", response_model=schemas.ContractorResponse)
async def update_contractor(cid: int, data: schemas.ContractorUpdate, db: Session = Depends(database.get_db)):
    c = db.query(models.Contractor).filter(models.Contractor.id == cid).first()
    if not c: raise HTTPException(404)
    for k, v in data.model_dump(exclude_unset=True).items(): setattr(c, k, v)
    db.commit(); db.refresh(c)
    return c


@app.delete("/contractors/{cid}")
async def delete_contractor(cid: int, db: Session = Depends(database.get_db)):
    c = db.query(models.Contractor).filter(models.Contractor.id == cid).first()
    if not c: raise HTTPException(404)
    db.delete(c); db.commit()
    return {"ok": True}


# =====================
# РАБОТЫ ПОДРЯДЧИКОВ (старые)
# =====================

@app.post("/contractor-works/")
async def create_work(data: schemas.ContractorWorkCreate, db: Session = Depends(database.get_db)):
    d = data.model_dump(); tool_ids = d.pop("tool_ids", [])
    if d.get("estimate_item_id"):
        item = db.query(models.EstimateItem).filter(models.EstimateItem.id == d["estimate_item_id"]).first()
        if item:
            if not d.get("quantity"): d["quantity"] = item.quantity
            if not d.get("unit"): d["unit"] = item.unit
            if not d.get("price_per_unit"): d["price_per_unit"] = round(item.price_per_unit * CONTRACTOR_MARGIN, 2)
            if not d.get("total_price"): d["total_price"] = round(d["quantity"] * d["price_per_unit"], 2)
            if not d.get("description"): d["description"] = item.name
    if not d.get("description"): raise HTTPException(400, "Укажите описание")
    work = models.ContractorWork(**d, created_at=now_str()); db.add(work); db.flush()
    if tool_ids: work.tools = db.query(models.Tool).filter(models.Tool.id.in_(tool_ids)).all()
    db.commit(); db.refresh(work)
    return build_work_response(work)


@app.get("/contractor-works/")
async def list_works(object_id: Optional[int] = None, contractor_id: Optional[int] = None,
                     status: Optional[str] = None, db: Session = Depends(database.get_db)):
    q = db.query(models.ContractorWork).options(
        joinedload(models.ContractorWork.contractor), joinedload(models.ContractorWork.tools))
    if object_id: q = q.filter(models.ContractorWork.object_id == object_id)
    if contractor_id: q = q.filter(models.ContractorWork.contractor_id == contractor_id)
    if status: q = q.filter(models.ContractorWork.status == status)
    return [build_work_response(w) for w in q.order_by(models.ContractorWork.id.desc()).all()]


@app.put("/contractor-works/{work_id}")
async def update_work(work_id: int, data: schemas.ContractorWorkUpdate, db: Session = Depends(database.get_db)):
    w = db.query(models.ContractorWork).filter(models.ContractorWork.id == work_id).first()
    if not w: raise HTTPException(404)
    d = data.model_dump(exclude_unset=True); tool_ids = d.pop("tool_ids", None)
    for k, v in d.items(): setattr(w, k, v)
    if tool_ids is not None: w.tools = db.query(models.Tool).filter(models.Tool.id.in_(tool_ids)).all()
    db.commit(); db.refresh(w)
    return build_work_response(w)


@app.delete("/contractor-works/{work_id}")
async def delete_work(work_id: int, db: Session = Depends(database.get_db)):
    w = db.query(models.ContractorWork).filter(models.ContractorWork.id == work_id).first()
    if not w: raise HTTPException(404)
    db.delete(w); db.commit()
    return {"ok": True}


# =====================
# СМЕТЫ ПОДРЯДЧИКА (с привязкой к позициям сметы объекта)
# =====================

@app.post("/contractor-estimates/")
async def create_contractor_estimate(data: schemas.ContractorEstimateCreate, db: Session = Depends(database.get_db)):
    obj = db.query(models.Object).filter(models.Object.id == data.object_id).first()
    if not obj: raise HTTPException(404, "Объект не найден")
    contr = db.query(models.Contractor).filter(models.Contractor.id == data.contractor_id).first()
    if not contr: raise HTTPException(404, "Подрядчик не найден")

    # ПРОВЕРКА: не заняты ли позиции другими подрядчиками (если статус блокирующий)
    if data.status in ("in_progress", "completed", "paid", "Акт"):
        locked_ids = set()
        existing_estimates = db.query(models.ContractorEstimate).filter(
            models.ContractorEstimate.object_id == data.object_id,
            models.ContractorEstimate.contractor_id != data.contractor_id,
            models.ContractorEstimate.status.in_(["in_progress", "completed", "paid", "Акт"])
        ).all()
        for est in existing_estimates:
            for it in est.items:
                if it.estimate_item_id:
                    locked_ids.add(it.estimate_item_id)
        for item_data in data.items:
            if item_data.estimate_item_id and item_data.estimate_item_id in locked_ids:
                raise HTTPException(400, f"Позиция '{item_data.name}' уже выполняется другим подрядчиком")

    name = data.name or f"Смета работ: {obj.name}"
    estimate = models.ContractorEstimate(
        object_id=data.object_id, contractor_id=data.contractor_id,
        name=name, status=data.status, notes=data.notes, created_at=now_str())
    db.add(estimate); db.flush()
    total = 0
    for item_data in data.items:
        item_total = round(item_data.quantity * item_data.price_per_unit, 2)
        db.add(models.ContractorEstimateItem(
            estimate_id=estimate.id,
            estimate_item_id=item_data.estimate_item_id,
            name=item_data.name, unit=item_data.unit,
            quantity=item_data.quantity, price_per_unit=item_data.price_per_unit,
            total_price=item_total))
        total += item_total
    estimate.total_sum = round(total, 2)
    db.commit(); db.refresh(estimate)
    # Reload with joins
    estimate = db.query(models.ContractorEstimate).options(
        joinedload(models.ContractorEstimate.contractor),
        joinedload(models.ContractorEstimate.items).joinedload(models.ContractorEstimateItem.estimate_item),
    ).filter(models.ContractorEstimate.id == estimate.id).first()
    return build_contractor_estimate_response(estimate)


@app.get("/contractor-estimates/")
async def list_contractor_estimates(object_id: Optional[int] = None, contractor_id: Optional[int] = None,
                                     status: Optional[str] = None, db: Session = Depends(database.get_db)):
    q = db.query(models.ContractorEstimate).options(
        joinedload(models.ContractorEstimate.contractor),
        joinedload(models.ContractorEstimate.items).joinedload(models.ContractorEstimateItem.estimate_item))
    if object_id: q = q.filter(models.ContractorEstimate.object_id == object_id)
    if contractor_id: q = q.filter(models.ContractorEstimate.contractor_id == contractor_id)
    if status: q = q.filter(models.ContractorEstimate.status == status)
    return [build_contractor_estimate_response(e) for e in q.order_by(models.ContractorEstimate.id.desc()).all()]


@app.get("/contractor-estimates/{eid}")
async def get_contractor_estimate(eid: int, db: Session = Depends(database.get_db)):
    e = db.query(models.ContractorEstimate).options(
        joinedload(models.ContractorEstimate.contractor),
        joinedload(models.ContractorEstimate.items).joinedload(models.ContractorEstimateItem.estimate_item),
    ).filter(models.ContractorEstimate.id == eid).first()
    if not e: raise HTTPException(404)
    return build_contractor_estimate_response(e)


@app.put("/contractor-estimates/{eid}")
async def update_contractor_estimate(eid: int, data: schemas.ContractorEstimateUpdate, db: Session = Depends(database.get_db)):
    e = db.query(models.ContractorEstimate).filter(models.ContractorEstimate.id == eid).first()
    if not e: raise HTTPException(404)
    update_data = data.model_dump(exclude_unset=True)
    items_data = update_data.pop("items", None)

    old_status = e.status
    for k, v in update_data.items(): setattr(e, k, v)

    # Track status transitions
    if "status" in update_data:
        if update_data["status"] == "completed" and old_status != "completed":
            e.completed_at = now_str()
        if update_data["status"] == "paid" and old_status != "paid":
            e.paid_at = now_str()

    if items_data is not None:
        db.query(models.ContractorEstimateItem).filter(
            models.ContractorEstimateItem.estimate_id == eid).delete()
        total = 0
        for item_raw in items_data:
            if isinstance(item_raw, dict):
                i_data = item_raw
            else:
                i_data = item_raw.model_dump() if hasattr(item_raw, 'model_dump') else dict(item_raw)
            i_total = round(float(i_data.get("quantity", 0)) * float(i_data.get("price_per_unit", 0)), 2)
            db.add(models.ContractorEstimateItem(
                estimate_id=eid,
                estimate_item_id=i_data.get("estimate_item_id"),
                name=i_data["name"], unit=i_data.get("unit", "шт"),
                quantity=i_data.get("quantity", 1), price_per_unit=i_data.get("price_per_unit", 0),
                total_price=i_total))
            total += i_total
        e.total_sum = round(total, 2)

    db.commit(); db.refresh(e)
    e = db.query(models.ContractorEstimate).options(
        joinedload(models.ContractorEstimate.contractor),
        joinedload(models.ContractorEstimate.items).joinedload(models.ContractorEstimateItem.estimate_item),
    ).filter(models.ContractorEstimate.id == eid).first()
    return build_contractor_estimate_response(e)


@app.delete("/contractor-estimates/{eid}")
async def delete_contractor_estimate(eid: int, db: Session = Depends(database.get_db)):
    e = db.query(models.ContractorEstimate).filter(models.ContractorEstimate.id == eid).first()
    if not e: raise HTTPException(404)
    db.delete(e); db.commit()
    return {"ok": True}


@app.get("/contractor-estimates/{eid}/excel")
async def download_estimate_excel(eid: int, db: Session = Depends(database.get_db)):
    e = db.query(models.ContractorEstimate).options(
        joinedload(models.ContractorEstimate.contractor),
        joinedload(models.ContractorEstimate.items).joinedload(models.ContractorEstimateItem.estimate_item),
    ).filter(models.ContractorEstimate.id == eid).first()
    if not e: raise HTTPException(404)
    wb = Workbook(); ws = wb.active; ws.title = "Смета работ"
    bold = Font(bold=True); center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    money_fmt = '#,##0.00" ₽"'
    hfill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    ws.merge_cells("A1:G1"); ws["A1"] = e.name; ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A2:G2")
    ws["A2"] = f"Подрядчик: {e.contractor.name if e.contractor else '—'}"
    col_headers = ["№", "Наименование", "Ед.", "Кол-во", "Цена подр.", "Сумма подр.", "Цена клиента"]
    for ci, h in enumerate(col_headers, 1):
        c = ws.cell(row=4, column=ci, value=h); c.font = bold; c.alignment = center; c.border = thin; c.fill = hfill
    row = 5
    for idx, item in enumerate(e.items, 1):
        ws.cell(row=row, column=1, value=idx).alignment = center
        ws.cell(row=row, column=2, value=item.name)
        ws.cell(row=row, column=3, value=item.unit).alignment = center
        ws.cell(row=row, column=4, value=item.quantity).alignment = center
        ws.cell(row=row, column=5, value=item.price_per_unit).number_format = money_fmt
        ws.cell(row=row, column=6, value=item.total_price).number_format = money_fmt
        client_total = item.estimate_item.total_price if item.estimate_item else 0
        ws.cell(row=row, column=7, value=client_total).number_format = money_fmt
        for col in range(1, 8): ws.cell(row=row, column=col).border = thin
        row += 1
    row += 1
    ws.merge_cells(f"A{row}:E{row}"); ws.cell(row=row, column=1, value="ИТОГО:").font = bold
    ws.cell(row=row, column=6, value=e.total_sum).font = bold
    ws.column_dimensions["A"].width = 5; ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 8; ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 14; ws.column_dimensions["F"].width = 14; ws.column_dimensions["G"].width = 14
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename=estimate_{e.id}.xlsx"})


# =====================
# ОПЛАТЫ ПОДРЯДЧИКАМ (НОВОЕ)
# =====================

@app.post("/contractor-payments/")
async def create_contractor_payment(data: schemas.ContractorPaymentCreate, db: Session = Depends(database.get_db)):
    payload = data.model_dump()
    if not payload.get("date"): payload["date"] = today_str()
    p = models.ContractorPayment(**payload); db.add(p); db.commit(); db.refresh(p)
    return {"id": p.id, "amount": p.amount, "contractor_id": p.contractor_id}


@app.get("/contractor-payments/")
async def list_contractor_payments(contractor_id: Optional[int] = None, db: Session = Depends(database.get_db)):
    q = db.query(models.ContractorPayment).options(joinedload(models.ContractorPayment.contractor))
    if contractor_id: q = q.filter(models.ContractorPayment.contractor_id == contractor_id)
    return [{"id": p.id, "contractor_id": p.contractor_id,
             "contractor_name": p.contractor.name if p.contractor else "",
             "object_id": p.object_id, "amount": p.amount,
             "description": p.description, "date": p.date}
            for p in q.order_by(models.ContractorPayment.id.desc()).all()]


@app.delete("/contractor-payments/{pid}")
async def delete_contractor_payment(pid: int, db: Session = Depends(database.get_db)):
    p = db.query(models.ContractorPayment).filter(models.ContractorPayment.id == pid).first()
    if not p: raise HTTPException(404)
    db.delete(p); db.commit()
    return {"ok": True}


# =====================
# ДОП. РАБОТЫ (НОВОЕ)
# =====================

@app.post("/extra-works/")
async def create_extra_work(data: schemas.ExtraWorkCreate, db: Session = Depends(database.get_db)):
    total = round(data.quantity * data.price, 2)
    w = models.ExtraWork(object_id=data.object_id, name=data.name,
                         description=data.description or "", contractor_name=data.contractor_name or "",
                         quantity=data.quantity, unit=data.unit or "", price=data.price,
                         total_price=total, status="completed", created_at=now_str())
    db.add(w); db.commit(); db.refresh(w)
    return {"id": w.id, "total_price": w.total_price}


@app.get("/extra-works/")
async def list_extra_works(object_id: Optional[int] = None, db: Session = Depends(database.get_db)):
    q = db.query(models.ExtraWork).options(joinedload(models.ExtraWork.object))
    if object_id: q = q.filter(models.ExtraWork.object_id == object_id)
    return [{"id": w.id, "object_id": w.object_id, "object_name": w.object.name if w.object else "",
             "name": w.name, "description": w.description, "contractor_name": w.contractor_name,
             "quantity": w.quantity, "unit": w.unit, "price": w.price,
             "total_price": w.total_price, "status": w.status,
             "created_at": w.created_at, "paid_at": w.paid_at}
            for w in q.order_by(models.ExtraWork.id.desc()).all()]


@app.put("/extra-works/{wid}")
async def update_extra_work(wid: int, data: schemas.ExtraWorkUpdate, db: Session = Depends(database.get_db)):
    w = db.query(models.ExtraWork).filter(models.ExtraWork.id == wid).first()
    if not w: raise HTTPException(404)
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(w, k, v)
    if data.status == "paid" and not w.paid_at: w.paid_at = now_str()
    if data.quantity is not None or data.price is not None:
        w.total_price = round(w.quantity * w.price, 2)
    db.commit(); db.refresh(w)
    return {"ok": True}


@app.delete("/extra-works/{wid}")
async def delete_extra_work(wid: int, db: Session = Depends(database.get_db)):
    w = db.query(models.ExtraWork).filter(models.ExtraWork.id == wid).first()
    if not w: raise HTTPException(404)
    db.delete(w); db.commit()
    return {"ok": True}


# =====================
# ВКЛАДКА «РАБОТЫ» — комбинированный эндпоинт
# =====================

@app.get("/works-combined/")
async def get_works_combined(object_id: Optional[int] = None, db: Session = Depends(database.get_db)):
    """Возвращает объединённый список: сметы подрядчика (completed/paid) + доп. работы"""
    result = []

    # 1. Сметы подрядчика со статусом completed или paid
    est_q = db.query(models.ContractorEstimate).options(
        joinedload(models.ContractorEstimate.contractor),
        joinedload(models.ContractorEstimate.items).joinedload(models.ContractorEstimateItem.estimate_item),
        joinedload(models.ContractorEstimate.object),
    ).filter(models.ContractorEstimate.status.in_(["completed", "paid"]))
    if object_id: est_q = est_q.filter(models.ContractorEstimate.object_id == object_id)

    for est in est_q.order_by(models.ContractorEstimate.id.desc()).all():
        client_total = sum(
            (i.estimate_item.total_price if i.estimate_item else i.total_price) for i in est.items
        )
        result.append({
            "id": f"est_{est.id}", "type": "estimate", "estimate_id": est.id,
            "object_id": est.object_id, "object_name": est.object.name if est.object else "",
            "name": est.name, "contractor_name": est.contractor.name if est.contractor else "",
            "status": est.status, "total_sum": est.total_sum, "client_total": round2(client_total),
            "created_at": est.created_at, "completed_at": est.completed_at, "paid_at": est.paid_at,
        })

    # 2. Доп. работы
    ew_q = db.query(models.ExtraWork).options(joinedload(models.ExtraWork.object))
    if object_id: ew_q = ew_q.filter(models.ExtraWork.object_id == object_id)

    for w in ew_q.order_by(models.ExtraWork.id.desc()).all():
        result.append({
            "id": f"work_{w.id}", "type": "extra", "work_id": w.id,
            "object_id": w.object_id, "object_name": w.object.name if w.object else "",
            "name": w.name, "contractor_name": w.contractor_name or "",
            "description": w.description, "quantity": w.quantity, "unit": w.unit,
            "price": w.price, "total_price": w.total_price, "status": w.status,
            "created_at": w.created_at, "paid_at": w.paid_at,
        })

    return result


# =====================
# МАТЕРИАЛЫ
# =====================

@app.post("/materials/", response_model=schemas.MaterialResponse)
async def create_material(data: schemas.MaterialCreate, db: Session = Depends(database.get_db)):
    m = models.Material(**data.model_dump()); db.add(m); db.commit(); db.refresh(m)
    return m


@app.get("/materials/", response_model=List[schemas.MaterialResponse])
async def list_materials(db: Session = Depends(database.get_db)):
    return db.query(models.Material).all()


@app.put("/materials/{mid}", response_model=schemas.MaterialResponse)
async def update_material(mid: int, data: schemas.MaterialUpdate, db: Session = Depends(database.get_db)):
    m = db.query(models.Material).filter(models.Material.id == mid).first()
    if not m: raise HTTPException(404)
    for k, v in data.model_dump(exclude_unset=True).items(): setattr(m, k, v)
    db.commit(); db.refresh(m)
    return m


@app.delete("/materials/{mid}")
async def delete_material(mid: int, db: Session = Depends(database.get_db)):
    m = db.query(models.Material).filter(models.Material.id == mid).first()
    if not m: raise HTTPException(404)
    db.delete(m); db.commit()
    return {"ok": True}


# =====================
# ЗАКУПКИ
# =====================

@app.post("/material-purchases/")
async def create_purchase(data: schemas.MaterialPurchaseCreate, db: Session = Depends(database.get_db)):
    payload = data.model_dump()
    if not payload.get("date"): payload["date"] = today_str()
    p = models.MaterialPurchase(**payload); db.add(p); db.commit(); db.refresh(p)
    _ = p.material
    return {"id": p.id, "total_price": p.total_price}


@app.get("/material-purchases/")
async def list_purchases(object_id: Optional[int] = None, db: Session = Depends(database.get_db)):
    q = db.query(models.MaterialPurchase).options(joinedload(models.MaterialPurchase.material))
    if object_id: q = q.filter(models.MaterialPurchase.object_id == object_id)
    return [{"id": p.id, "object_id": p.object_id, "material_id": p.material_id,
             "quantity": p.quantity, "total_price": p.total_price,
             "supplier": p.supplier, "date": p.date, "status": p.status, "notes": p.notes,
             "material": {"id": p.material.id, "name": p.material.name, "unit": p.material.unit,
                          "price_per_unit": p.material.price_per_unit,
                          "description": p.material.description} if p.material else None}
            for p in q.order_by(models.MaterialPurchase.id.desc()).all()]


@app.put("/material-purchases/{pid}")
async def update_purchase(pid: int, data: schemas.MaterialPurchaseUpdate, db: Session = Depends(database.get_db)):
    p = db.query(models.MaterialPurchase).filter(models.MaterialPurchase.id == pid).first()
    if not p: raise HTTPException(404)
    for k, v in data.model_dump(exclude_unset=True).items(): setattr(p, k, v)
    db.commit()
    return {"ok": True}


@app.delete("/material-purchases/{pid}")
async def delete_purchase(pid: int, db: Session = Depends(database.get_db)):
    p = db.query(models.MaterialPurchase).filter(models.MaterialPurchase.id == pid).first()
    if not p: raise HTTPException(404)
    db.delete(p); db.commit()
    return {"ok": True}


# =====================
# ИНСТРУМЕНТЫ
# =====================

@app.post("/tools/", response_model=schemas.ToolResponse)
async def create_tool(data: schemas.ToolCreate, db: Session = Depends(database.get_db)):
    t = models.Tool(**data.model_dump()); db.add(t); db.commit(); db.refresh(t)
    return t


@app.get("/tools/", response_model=List[schemas.ToolResponse])
async def list_tools(contractor_id: Optional[int] = None, status: Optional[str] = None,
                     db: Session = Depends(database.get_db)):
    q = db.query(models.Tool)
    if contractor_id: q = q.filter(models.Tool.contractor_id == contractor_id)
    if status: q = q.filter(models.Tool.status == status)
    return q.all()


@app.put("/tools/{tid}", response_model=schemas.ToolResponse)
async def update_tool(tid: int, data: schemas.ToolUpdate, db: Session = Depends(database.get_db)):
    t = db.query(models.Tool).filter(models.Tool.id == tid).first()
    if not t: raise HTTPException(404)
    for k, v in data.model_dump(exclude_unset=True).items(): setattr(t, k, v)
    db.commit(); db.refresh(t)
    return t


@app.delete("/tools/{tid}")
async def delete_tool(tid: int, db: Session = Depends(database.get_db)):
    t = db.query(models.Tool).filter(models.Tool.id == tid).first()
    if not t: raise HTTPException(404)
    db.delete(t); db.commit()
    return {"ok": True}


# =====================
# ЗАДАЧИ
# =====================

@app.post("/tasks/")
async def create_task(data: schemas.TaskCreate, db: Session = Depends(database.get_db)):
    t = models.Task(**data.model_dump(), created_at=now_str()); db.add(t); db.commit(); db.refresh(t)
    return t


@app.get("/tasks/")
async def list_tasks(object_id: Optional[int] = None, db: Session = Depends(database.get_db)):
    q = db.query(models.Task)
    if object_id: q = q.filter(models.Task.object_id == object_id)
    return q.all()


@app.put("/tasks/{tid}")
async def update_task(tid: int, data: schemas.TaskUpdate, db: Session = Depends(database.get_db)):
    t = db.query(models.Task).filter(models.Task.id == tid).first()
    if not t: raise HTTPException(404)
    for k, v in data.model_dump(exclude_unset=True).items(): setattr(t, k, v)
    db.commit(); db.refresh(t)
    return t


@app.delete("/tasks/{tid}")
async def delete_task(tid: int, db: Session = Depends(database.get_db)):
    t = db.query(models.Task).filter(models.Task.id == tid).first()
    if not t: raise HTTPException(404)
    db.delete(t); db.commit()
    return {"ok": True}


# =====================
# КОММУНИКАЦИИ
# =====================

@app.post("/communications/")
async def create_comm(data: schemas.CommunicationCreate, db: Session = Depends(database.get_db)):
    c = models.Communication(**data.model_dump()); db.add(c); db.commit(); db.refresh(c)
    return c


# =====================
# ОБЩИЕ РАСХОДЫ
# =====================

@app.post("/general-expenses/")
async def create_general_expense(data: schemas.GeneralExpenseCreate, db: Session = Depends(database.get_db)):
    payload = data.model_dump()
    if not payload.get("date"): payload["date"] = today_str()
    e = models.GeneralExpense(**payload); db.add(e); db.commit(); db.refresh(e)
    return e


@app.get("/general-expenses/")
async def list_general_expenses(db: Session = Depends(database.get_db)):
    return db.query(models.GeneralExpense).order_by(models.GeneralExpense.id.desc()).all()


@app.delete("/general-expenses/{eid}")
async def delete_general_expense(eid: int, db: Session = Depends(database.get_db)):
    e = db.query(models.GeneralExpense).filter(models.GeneralExpense.id == eid).first()
    if not e: raise HTTPException(404)
    db.delete(e); db.commit()
    return {"ok": True}


# =====================
# АКТЫ ВЫПОЛНЕННЫХ РАБОТ
# =====================

def _generate_act_number(db: Session) -> str:
    """Генерация номера акта: АКТ-001, АКТ-002 и т.д."""
    last = db.query(models.Act).order_by(models.Act.id.desc()).first()
    if last:
        try:
            num = int(last.act_number.split("-")[1]) + 1
        except (IndexError, ValueError):
            num = 1
    else:
        num = 1
    return f"АКТ-{num:03d}"


def _collect_act_items(db: Session, object_id: int) -> dict:
    """Сбор позиций акта из выполненных смет и доп. работ (клиентские цены).
    Возвращает dict: items (список позиций), estimate_ids (ID смет), extra_work_ids (ID доп. работ)."""
    items = []
    estimate_ids = []
    extra_work_ids = []

    # Собираем клиентские цены из смет объекта (по наименованию)
    client_prices = {}  # name -> {price, total_price, unit}
    estimates = db.query(models.Estimate).filter(models.Estimate.object_id == object_id).all()
    for est in estimates:
        for item in est.items:
            client_prices[item.name] = {
                "unit": item.unit,
                "price": item.price_per_unit,
                "total_price": item.total_price,
            }

    # Сметы подрядчика со статусом completed
    contractor_estimates = db.query(models.ContractorEstimate).filter(
        models.ContractorEstimate.object_id == object_id,
        models.ContractorEstimate.status == "completed"
    ).all()
    for est in contractor_estimates:
        estimate_ids.append(est.id)
        for item in est.items:
            # Ищем клиентскую цену по наименованию
            cp = client_prices.get(item.name)
            if cp:
                items.append({
                    "name": item.name,
                    "unit": cp["unit"],
                    "quantity": item.quantity,
                    "price_per_unit": cp["price"],
                    "total_price": cp["total_price"],
                    "source": "estimate"
                })
            else:
                # Нет клиентской цены — берём из подрядчика
                items.append({
                    "name": item.name,
                    "unit": item.unit,
                    "quantity": item.quantity,
                    "price_per_unit": item.price_per_unit,
                    "total_price": item.total_price,
                    "source": "estimate"
                })

    # Доп. работы со статусом completed
    extra_works = db.query(models.ExtraWork).filter(
        models.ExtraWork.object_id == object_id,
        models.ExtraWork.status == "completed"
    ).all()
    for w in extra_works:
        extra_work_ids.append(w.id)
        items.append({
            "name": w.name,
            "unit": w.unit,
            "quantity": w.quantity,
            "price_per_unit": w.price,
            "total_price": w.total_price,
            "source": "extra_work"
        })
    return {"items": items, "estimate_ids": estimate_ids, "extra_work_ids": extra_work_ids}


@app.post("/acts/")
async def create_act(data: schemas.ActCreate, db: Session = Depends(database.get_db)):
    """Создать акт из выполненных работ."""
    obj = db.query(models.Object).filter(models.Object.id == data.object_id).first()
    if not obj:
        raise HTTPException(404, "Объект не найден")

    collected = _collect_act_items(db, data.object_id)
    items = collected["items"]
    if not items:
        raise HTTPException(400, "Нет выполненных работ для формирования акта")

    total_sum = sum(i["total_price"] for i in items)
    act = models.Act(
        object_id=data.object_id,
        act_number=_generate_act_number(db),
        created_at=today_str(),
        total_sum=total_sum,
        notes=data.notes or ""
    )
    db.add(act)
    db.commit()
    db.refresh(act)

    # Меняем статусы смет и доп. работ на "Акт"
    for est_id in collected["estimate_ids"]:
        db.query(models.ContractorEstimate).filter(
            models.ContractorEstimate.id == est_id
        ).update({"status": "Акт"})
    for ew_id in collected["extra_work_ids"]:
        db.query(models.ExtraWork).filter(
            models.ExtraWork.id == ew_id
        ).update({"status": "Акт"})
    db.commit()

    return {"id": act.id, "act_number": act.act_number, "total_sum": act.total_sum, "items_count": len(items)}


@app.get("/acts/")
async def list_acts(object_id: Optional[int] = None, db: Session = Depends(database.get_db)):
    """Список актов."""
    q = db.query(models.Act).options(joinedload(models.Act.object)).order_by(models.Act.id.desc())
    if object_id:
        q = q.filter(models.Act.object_id == object_id)
    result = []
    for act in q.all():
        result.append({
            "id": act.id,
            "act_number": act.act_number,
            "object_id": act.object_id,
            "object_name": act.object.name if act.object else "",
            "created_at": act.created_at,
            "status": act.status,
            "total_sum": act.total_sum,
            "signed_at": act.signed_at
        })
    return result


@app.get("/acts/{act_id}")
async def get_act(act_id: int, db: Session = Depends(database.get_db)):
    """Детали акта с позициями."""
    act = db.query(models.Act).filter(models.Act.id == act_id).first()
    if not act:
        raise HTTPException(404, "Акт не найден")

    items = _collect_act_items(db, act.object_id)["items"]

    return {
        "id": act.id,
        "object_id": act.object_id,
        "object_name": act.object.name if act.object else "",
        "act_number": act.act_number,
        "created_at": act.created_at,
        "status": act.status,
        "total_sum": act.total_sum,
        "notes": act.notes,
        "signed_at": act.signed_at,
        "items": items
    }


@app.put("/acts/{act_id}")
async def update_act(act_id: int, data: schemas.ActUpdate, db: Session = Depends(database.get_db)):
    """Обновить акт (статус, заметки)."""
    act = db.query(models.Act).filter(models.Act.id == act_id).first()
    if not act:
        raise HTTPException(404, "Акт не найден")
    
    if data.status is not None:
        act.status = data.status
        if data.status == "Оплачен" and not act.signed_at:
            act.signed_at = today_str()
    if data.notes is not None:
        act.notes = data.notes
    
    db.commit()
    db.refresh(act)
    return {"id": act.id, "status": act.status, "notes": act.notes}


@app.delete("/acts/{act_id}")
async def delete_act(act_id: int, db: Session = Depends(database.get_db)):
    """Удалить акт."""
    act = db.query(models.Act).filter(models.Act.id == act_id).first()
    if not act:
        raise HTTPException(404, "Акт не найден")
    db.delete(act)
    db.commit()
    return {"ok": True}


@app.get("/acts/{act_id}/excel")
async def export_act(act_id: int, db: Session = Depends(database.get_db)):
    """Экспорт акта в Excel."""
    act = db.query(models.Act).filter(models.Act.id == act_id).first()
    if not act:
        raise HTTPException(404, "Акт не найден")

    items = _collect_act_items(db, act.object_id)["items"]
    obj = db.query(models.Object).filter(models.Object.id == act.object_id).first()

    wb = Workbook(); ws = wb.active; ws.title = "Акт выполненных работ"
    bold = Font(bold=True); center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    money_fmt = '#,##0.00" ₽"'
    hfill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    ws.merge_cells("A1:F1"); ws["A1"] = f"АКТ ВЫПОЛНЕННЫХ РАБОТ {act.act_number}"; ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A2:F2"); ws["A2"] = f"Объект: {obj.name if obj else '—'} | Дата: {act.created_at} | Статус: {act.status}"

    col_headers = ["№", "Наименование", "Ед.", "Кол-во", "Цена", "Сумма"]
    for ci, h in enumerate(col_headers, 1):
        c = ws.cell(row=4, column=ci, value=h); c.font = bold; c.alignment = center; c.border = thin; c.fill = hfill

    row = 5
    for idx, item in enumerate(items, 1):
        ws.cell(row=row, column=1, value=idx).alignment = center
        ws.cell(row=row, column=2, value=item["name"])
        ws.cell(row=row, column=3, value=item["unit"]).alignment = center
        ws.cell(row=row, column=4, value=item["quantity"]).alignment = center
        ws.cell(row=row, column=5, value=item["price_per_unit"]).number_format = money_fmt
        ws.cell(row=row, column=6, value=item["total_price"]).number_format = money_fmt
        for col in range(1, 7): ws.cell(row=row, column=col).border = thin
        row += 1

    row += 1
    ws.merge_cells(f"A{row}:D{row}"); ws.cell(row=row, column=1, value="ИТОГО:").font = bold
    ws.cell(row=row, column=6, value=act.total_sum).font = bold
    ws.cell(row=row, column=6).number_format = money_fmt

    # Подписи
    row += 3
    ws.cell(row=row, column=1, value="Заказчик: ________________ / ________________")
    ws.cell(row=row+1, column=1, value="Подрядчик: ________________ / ________________")

    ws.column_dimensions["A"].width = 5; ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 8; ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 14; ws.column_dimensions["F"].width = 14

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    from urllib.parse import quote
    filename = quote(f"{act.act_number}.xlsx", safe='')
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"})


# =====================
# ФИНАНСЫ (ПЕРЕРАБОТАНО)
# =====================

@app.get("/financial/summary")
async def financial_summary(db: Session = Depends(database.get_db)):
    # ДОХОДЫ: платежи (paid) + акты (Оплачен)
    total_debit_payments = db.query(func.sum(models.Payment.amount)).filter(
        models.Payment.status == "paid").scalar() or 0
    total_debit_acts = db.query(func.sum(models.Act.total_sum)).filter(
        models.Act.status == "Оплачен").scalar() or 0
    total_debit = float(total_debit_payments) + float(total_debit_acts)

    # РАСХОДЫ: сметы подрядчика (paid) + доп. работы (paid) + старые работы + материалы + общие
    total_contr_est = db.query(func.sum(models.ContractorEstimate.total_sum)).filter(
        models.ContractorEstimate.status == "paid").scalar() or 0
    total_extra_work = db.query(func.sum(models.ExtraWork.total_price)).filter(
        models.ExtraWork.status == "paid").scalar() or 0
    total_old_work = db.query(func.sum(models.ContractorWork.total_price)).filter(
        models.ContractorWork.status.in_(["completed", "paid"])).scalar() or 0
    total_contractor = float(total_contr_est) + float(total_extra_work) + float(total_old_work)

    total_mat_req = db.query(func.sum(models.MaterialRequest.total_sum)).filter(
        models.MaterialRequest.status == "Оплачено").scalar() or 0
    total_mat_purch = db.query(func.sum(models.MaterialPurchase.total_price)).filter(
        models.MaterialPurchase.status == "delivered").scalar() or 0
    total_material = float(total_mat_req) + float(total_mat_purch)

    total_general = db.query(func.sum(models.GeneralExpense.amount)).scalar() or 0
    total_credit = total_contractor + total_material + float(total_general)

    # ПО ОБЪЕКТАМ
    objects = db.query(models.Object).options(*_object_joinedload_options()).all()
    by_objects = []
    for obj in objects:
        fin = calc_object_financials_from_loaded(obj)
        if fin["total_debit"] > 0 or fin["total_credit"] > 0:
            by_objects.append({"object_id": obj.id, "object_name": obj.name,
                               "client_name": obj.client_name, **fin})

    # ДОЛГИ КЛИЕНТОВ (НОВАЯ ЛОГИКА)
    client_debts = []
    for obj in objects:
        client_owed = 0
        # Выполненные позиции из смет подрядчика → по клиентским ценам
        for est in obj.contractor_estimates:
            if est.status in ("completed", "paid"):
                for item in est.items:
                    if item.estimate_item:
                        client_owed += item.estimate_item.total_price  # Клиентская цена
                    else:
                        client_owed += item.total_price  # Fallback
        # Доп. работы
        for w in obj.extra_works:
            if w.status in ("completed", "paid"):
                client_owed += w.total_price
        # Материалы (оплаченные заявки + закупки)
        for mr in obj.material_requests:
            if mr.status == "Оплачено":
                client_owed += mr.total_sum
        for mp in obj.material_purchases:
            if mp.status == "delivered":
                client_owed += mp.total_price

        client_paid = sum(p.amount for p in obj.payments if p.status == "paid")
        debt = client_owed - client_paid
        if debt > 0:
            client_debts.append({"object_id": obj.id, "object_name": obj.name,
                                 "client_name": obj.client_name, "debt": round2(debt)})

    # ДОЛГИ ПОДРЯДЧИКАМ (ИСПРАВЛЕННАЯ ЛОГИКА)
    contractors = db.query(models.Contractor).options(
        joinedload(models.Contractor.estimates),
        joinedload(models.Contractor.contractor_payments),
    ).all()
    contractor_debts = []
    for c in contractors:
        # Только выполненные работы (completed) - это реальный долг
        # paid - уже оплачены, не считаем
        # in_progress/planned - ещё не выполнены, не долг
        total_owed = sum(est.total_sum for est in c.estimates if est.status == "completed")
        # Доп. работы со статусом completed
        if hasattr(c, 'works'):
            total_owed += sum(w.total_price for w in c.works if w.status == "completed")
        # Все оплаты подрядчику
        total_paid_to = sum(p.amount for p in c.contractor_payments)
        debt = total_owed - total_paid_to
        if debt > 0:
            contractor_debts.append({"contractor_id": c.id, "name": c.name, "debt": round2(debt)})

    # ПЛАНИРУЕМЫЕ ДОХОДЫ (pending платежи + акты Выставлен)
    planned_income_payments = db.query(func.sum(models.Payment.amount)).filter(
        models.Payment.status == "pending").scalar() or 0
    planned_income_acts = db.query(func.sum(models.Act.total_sum)).filter(
        models.Act.status == "Выставлен").scalar() or 0
    planned_income = float(planned_income_payments) + float(planned_income_acts)

    # ПЛАНИРУЕМЫЕ РАСХОДЫ (in_progress и paid сметы + материалы в работе)
    planned_expenses_contractor = db.query(func.sum(models.ContractorEstimate.total_sum)).filter(
        models.ContractorEstimate.status.in_(["in_progress", "paid"])).scalar() or 0
    planned_expenses_materials = db.query(func.sum(models.MaterialRequest.total_sum)).filter(
        models.MaterialRequest.status.in_(["Отправлена", "Согласована", "Заказана"])).scalar() or 0
    planned_expenses = float(planned_expenses_contractor) + float(planned_expenses_materials)

    # ПЛАН
    cm = date.today().strftime("%Y-%m")
    plan = db.query(models.FinancialPlan).filter(models.FinancialPlan.month == cm).first()

    # Если план не задан (нулевой), используем авторасчёт
    use_auto = not plan or (plan.planned_income == 0 and plan.planned_expenses == 0)
    
    return {
        "total_debit": round2(total_debit),
        "total_contractor": round2(total_contractor),
        "total_material": round2(total_material),
        "total_general": round2(total_general),
        "total_credit": round2(total_credit),
        "total_profit": round2(float(total_debit) - total_credit),
        "by_objects": by_objects,
        "planned_income": round2(planned_income) if use_auto else plan.planned_income,
        "planned_expenses": round2(planned_expenses) if use_auto else plan.planned_expenses,
        "planned_profit": (round2(planned_income) - round2(planned_expenses)) if use_auto else (plan.planned_income - plan.planned_expenses),
        "contractor_debts": contractor_debts,
        "client_debts": client_debts,
    }


@app.post("/financial/plans/")
async def save_plan(data: schemas.FinancialPlanCreate, db: Session = Depends(database.get_db)):
    existing = db.query(models.FinancialPlan).filter(models.FinancialPlan.month == data.month).first()
    if existing:
        existing.planned_income = data.planned_income
        existing.planned_expenses = data.planned_expenses
        existing.notes = data.notes
        db.commit(); db.refresh(existing)
        return existing
    plan = models.FinancialPlan(**data.model_dump()); db.add(plan); db.commit(); db.refresh(plan)
    return plan


# =====================
# ДАШБОРД
# =====================

@app.get("/dashboard/")
async def dashboard(db: Session = Depends(database.get_db)):
    current_date = today_str()
    def _tq(): return db.query(models.Task).options(joinedload(models.Task.object))
    overdue = _tq().filter(models.Task.deadline < current_date, models.Task.status != "Done").all()
    today_tasks = _tq().filter(models.Task.deadline == current_date, models.Task.status != "Done").all()
    week_end = (date.today() + timedelta(days=7)).isoformat()
    week_tasks = _tq().filter(models.Task.deadline >= current_date, models.Task.deadline <= week_end,
                              models.Task.status != "Done").order_by(models.Task.deadline).limit(10).all()
    status_counts = db.query(models.Object.status, func.count(models.Object.id)).group_by(models.Object.status).all()
    pending = db.query(func.sum(models.Payment.amount)).filter(models.Payment.status == "pending").scalar() or 0
    paid = db.query(func.sum(models.Payment.amount)).filter(models.Payment.status == "paid").scalar() or 0
    active = db.query(models.Object).filter(models.Object.status != "Готово").limit(5).all()
    def td(t): return {"id": t.id, "title": t.title, "deadline": t.deadline,
                       "object_name": t.object.name if t.object else ""}
    return {
        "today_tasks": [td(t) for t in today_tasks],
        "overdue_tasks": [td(t) for t in overdue],
        "week_tasks": [td(t) for t in week_tasks],
        "status_counts": {s: c for s, c in status_counts},
        "finance": {"pending": round2(pending), "paid": round2(paid)},
        "active_objects": [{"id": o.id, "name": o.name, "client_name": o.client_name, "status": o.status}
                           for o in active],
    }


# =====================
# ПОСТАВЩИКИ
# =====================

@app.get("/suppliers/")
def list_suppliers(db: Session = Depends(database.get_db)):
    suppliers_list = db.query(models.Supplier).order_by(models.Supplier.id.desc()).all()
    result = []
    for s in suppliers_list:
        files = [{"id": f.id, "file_name": f.file_name, "file_path": f.file_path,
                  "description": f.description, "uploaded_at": f.uploaded_at,
                  "items_count": len(f.items),
                  "items": [{"id": it.id, "name": it.name, "unit": it.unit,
                             "price_per_unit": it.price_per_unit, "row_number": it.row_number}
                            for it in f.items]}
                 for f in s.price_files]
        result.append({"id": s.id, "name": s.name, "phone": s.phone, "email": s.email,
                       "address": s.address, "notes": s.notes, "created_at": s.created_at,
                       "price_files": files,
                       "total_items": sum(len(f.items) for f in s.price_files)})
    return result


@app.post("/suppliers/")
def create_supplier(data: dict, db: Session = Depends(database.get_db)):
    s = models.Supplier(name=data["name"], phone=data.get("phone", ""), email=data.get("email", ""),
                        address=data.get("address", ""), notes=data.get("notes", ""), created_at=now_str())
    db.add(s); db.commit(); db.refresh(s)
    return {"id": s.id, "name": s.name}


@app.put("/suppliers/{sid}")
def update_supplier(sid: int, data: dict, db: Session = Depends(database.get_db)):
    s = db.query(models.Supplier).filter(models.Supplier.id == sid).first()
    if not s: raise HTTPException(404)
    for f in ["name", "phone", "email", "address", "notes"]:
        if f in data: setattr(s, f, data[f])
    db.commit()
    return {"ok": True}


@app.delete("/suppliers/{sid}")
def delete_supplier(sid: int, db: Session = Depends(database.get_db)):
    s = db.query(models.Supplier).filter(models.Supplier.id == sid).first()
    if not s: raise HTTPException(404)
    for f in s.price_files:
        if f.file_path and os.path.exists(f.file_path): os.remove(f.file_path)
    db.delete(s); db.commit()
    return Response(status_code=204)


@app.post("/suppliers/{sid}/price-file/upload")
async def upload_supplier_price_file(sid: int, file: UploadFile = File(...),
                                      description: str = "", db: Session = Depends(database.get_db)):
    s = db.query(models.Supplier).filter(models.Supplier.id == sid).first()
    if not s: raise HTTPException(404)
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in [".xlsx", ".xls"]: raise HTTPException(400, "Только Excel")
    safe_name = f"supplier_{sid}_{int(time.time())}_{file.filename}"
    file_path = os.path.join(UPLOAD_DIR, safe_name)
    content = await file.read()
    with open(file_path, "wb") as f: f.write(content)
    items = parse_supplier_price_excel(content, file.filename)
    pf = models.SupplierPriceFile(supplier_id=sid, file_name=file.filename, file_path=file_path,
                                   description=description, uploaded_at=now_str())
    db.add(pf); db.flush()
    for item in items:
        db.add(models.SupplierPriceItem(price_file_id=pf.id, name=item["name"],
                                         unit=item.get("unit", "шт"),
                                         price_per_unit=item.get("price_per_unit", 0),
                                         row_number=item.get("row_number")))
    db.commit()
    return {"message": f"Загружено {len(items)} позиций", "file_id": pf.id}


@app.delete("/suppliers/price-files/{fid}")
def delete_supplier_price_file(fid: int, db: Session = Depends(database.get_db)):
    pf = db.query(models.SupplierPriceFile).filter(models.SupplierPriceFile.id == fid).first()
    if not pf: raise HTTPException(404)
    if pf.file_path and os.path.exists(pf.file_path): os.remove(pf.file_path)
    db.delete(pf); db.commit()
    return Response(status_code=204)


@app.get("/suppliers/{sid}/price-items/search")
def search_supplier_price_items(sid: int, q: str = "", db: Session = Depends(database.get_db)):
    if len(q) < 2: return []
    items = db.query(models.SupplierPriceItem).join(models.SupplierPriceFile).filter(
        models.SupplierPriceFile.supplier_id == sid,
        models.SupplierPriceItem.name.ilike(f"%{escape_like(q)}%")).limit(20).all()
    return [{"id": i.id, "name": i.name, "unit": i.unit, "price_per_unit": i.price_per_unit,
             "row_number": i.row_number} for i in items]


# =====================
# СПРАВОЧНИК МАТЕРИАЛОВ (НОВОЕ)
# =====================

@app.get("/supplier-catalog/")
def get_supplier_catalog(search: str = "", supplier_id: Optional[int] = None,
                          db: Session = Depends(database.get_db)):
    q = db.query(models.SupplierPriceItem, models.SupplierPriceFile, models.Supplier).join(
        models.SupplierPriceFile, models.SupplierPriceItem.price_file_id == models.SupplierPriceFile.id
    ).join(models.Supplier, models.SupplierPriceFile.supplier_id == models.Supplier.id)
    if search: q = q.filter(models.SupplierPriceItem.name.ilike(f"%{escape_like(search)}%"))
    if supplier_id: q = q.filter(models.Supplier.id == supplier_id)
    results = q.order_by(models.SupplierPriceItem.name).all()
    items = [{"id": item.id, "name": item.name, "unit": item.unit, "price_per_unit": item.price_per_unit,
              "row_number": item.row_number, "supplier_id": sup.id, "supplier_name": sup.name,
              "file_name": pf.file_name}
             for item, pf, sup in results]
    supplier_ids = set(i["supplier_id"] for i in items)
    return {"items": items, "total_items": len(items), "total_suppliers": len(supplier_ids)}


@app.get("/supplier-catalog/export")
def export_supplier_catalog(search: str = "", supplier_id: Optional[int] = None,
                             db: Session = Depends(database.get_db)):
    q = db.query(models.SupplierPriceItem, models.SupplierPriceFile, models.Supplier).join(
        models.SupplierPriceFile, models.SupplierPriceItem.price_file_id == models.SupplierPriceFile.id
    ).join(models.Supplier, models.SupplierPriceFile.supplier_id == models.Supplier.id)
    if search: q = q.filter(models.SupplierPriceItem.name.ilike(f"%{escape_like(search)}%"))
    if supplier_id: q = q.filter(models.Supplier.id == supplier_id)
    results = q.order_by(models.SupplierPriceItem.name).all()
    wb = Workbook(); ws = wb.active; ws.title = "Справочник материалов"
    headers = ["№", "Наименование", "Ед. изм.", "Цена за ед.", "Поставщик", "Файл"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h).font = Font(bold=True)
    for idx, (item, pf, sup) in enumerate(results, 1):
        ws.cell(row=idx + 1, column=1, value=idx)
        ws.cell(row=idx + 1, column=2, value=item.name)
        ws.cell(row=idx + 1, column=3, value=item.unit)
        ws.cell(row=idx + 1, column=4, value=item.price_per_unit)
        ws.cell(row=idx + 1, column=5, value=sup.name)
        ws.cell(row=idx + 1, column=6, value=pf.file_name)
    ws.column_dimensions["B"].width = 50; ws.column_dimensions["E"].width = 20
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=catalog_materials.xlsx"})


# =====================
# ЗАЯВКИ НА МАТЕРИАЛЫ
# =====================

@app.get("/material-requests/")
def list_material_requests(object_id: int = None, db: Session = Depends(database.get_db)):
    q = db.query(models.MaterialRequest).order_by(models.MaterialRequest.id.desc())
    if object_id: q = q.filter(models.MaterialRequest.object_id == object_id)
    result = []
    for r in q.all():
        obj = db.query(models.Object).filter(models.Object.id == r.object_id).first() if r.object_id else None
        sup = db.query(models.Supplier).filter(models.Supplier.id == r.supplier_id).first() if r.supplier_id else None
        result.append({
            "id": r.id, "name": r.name, "status": r.status, "total_sum": r.total_sum,
            "notes": r.notes, "created_at": r.created_at, "paid_at": r.paid_at, "file_path": r.file_path,
            "object_id": r.object_id, "object_name": obj.name if obj else "",
            "supplier_id": r.supplier_id, "supplier_name": sup.name if sup else "",
            "items": [{"id": it.id, "name": it.name, "unit": it.unit, "quantity": it.quantity,
                       "price_per_unit": it.price_per_unit, "total_price": it.total_price,
                       "supplier_price_item_id": it.supplier_price_item_id} for it in r.items]
        })
    return result


@app.post("/material-requests/")
def create_material_request(data: dict, db: Session = Depends(database.get_db)):
    mr = models.MaterialRequest(
        object_id=data.get("object_id"), supplier_id=data.get("supplier_id"),
        name=data.get("name", "Заявка"), status=data.get("status", "Черновик"),
        notes=data.get("notes", ""), created_at=now_str())
    db.add(mr); db.flush()
    total = 0
    for item in data.get("items", []):
        name = item.get("name", "").strip()
        if not name: continue
        qty = float(item.get("quantity", 1) or 1)
        price = float(item.get("price_per_unit", 0) or 0)
        tp = round(qty * price, 2); total += tp
        db.add(models.MaterialRequestItem(
            request_id=mr.id, name=name, unit=item.get("unit", "шт"),
            quantity=qty, price_per_unit=price, total_price=tp,
            supplier_price_item_id=item.get("supplier_price_item_id")))
    mr.total_sum = round(total, 2); db.commit(); db.refresh(mr)
    return {"id": mr.id, "total_sum": mr.total_sum}


@app.put("/material-requests/{rid}")
def update_material_request(rid: int, data: dict, db: Session = Depends(database.get_db)):
    mr = db.query(models.MaterialRequest).filter(models.MaterialRequest.id == rid).first()
    if not mr: raise HTTPException(404)
    for field in ["name", "status", "notes", "object_id", "supplier_id"]:
        if field in data: setattr(mr, field, data[field])
    if data.get("status") == "Оплачено" and not mr.paid_at:
        mr.paid_at = now_str()
    if "items" in data:
        db.query(models.MaterialRequestItem).filter(models.MaterialRequestItem.request_id == rid).delete()
        total = 0
        for item in data["items"]:
            name = item.get("name", "").strip()
            if not name: continue
            qty = float(item.get("quantity", 1) or 1)
            price = float(item.get("price_per_unit", 0) or 0)
            tp = round(qty * price, 2); total += tp
            db.add(models.MaterialRequestItem(
                request_id=mr.id, name=name, unit=item.get("unit", "шт"),
                quantity=qty, price_per_unit=price, total_price=tp,
                supplier_price_item_id=item.get("supplier_price_item_id")))
        mr.total_sum = round(total, 2)
    db.commit()
    return {"ok": True}


@app.delete("/material-requests/{rid}")
def delete_material_request(rid: int, db: Session = Depends(database.get_db)):
    mr = db.query(models.MaterialRequest).filter(models.MaterialRequest.id == rid).first()
    if not mr: raise HTTPException(404)
    if mr.file_path and os.path.exists(mr.file_path): os.remove(mr.file_path)
    db.delete(mr); db.commit()
    return Response(status_code=204)


@app.get("/material-requests/{rid}/export")
def export_material_request(rid: int, db: Session = Depends(database.get_db)):
    mr = db.query(models.MaterialRequest).filter(models.MaterialRequest.id == rid).first()
    if not mr: raise HTTPException(404)
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Заявка"
    thin = Side(style='thin'); border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.merge_cells('A1:F1'); ws['A1'] = mr.name; ws['A1'].font = Font(bold=True, size=14)
    obj = db.query(models.Object).filter(models.Object.id == mr.object_id).first() if mr.object_id else None
    sup = db.query(models.Supplier).filter(models.Supplier.id == mr.supplier_id).first() if mr.supplier_id else None
    row = 2
    if obj: ws.merge_cells(f'A{row}:F{row}'); ws[f'A{row}'] = f"Объект: {obj.name}"; row += 1
    if sup: ws.merge_cells(f'A{row}:F{row}'); ws[f'A{row}'] = f"Поставщик: {sup.name}"; row += 1
    row += 1
    headers = ["№", "Наименование", "Ед.", "Кол-во", "Цена", "Сумма"]
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=i+1, value=h); c.font = Font(bold=True, color="FFFFFF"); c.fill = hfill; c.border = border
    for idx, item in enumerate(mr.items, 1):
        row += 1
        for i, v in enumerate([idx, item.name, item.unit, item.quantity, item.price_per_unit, item.total_price]):
            c = ws.cell(row=row, column=i+1, value=v); c.border = border
    row += 1
    ws.merge_cells(f'A{row}:E{row}'); ws[f'A{row}'] = "ИТОГО:"; ws[f'A{row}'].font = Font(bold=True)
    ws.cell(row=row, column=6, value=mr.total_sum).font = Font(bold=True)
    ws.column_dimensions["B"].width = 45
    file_name = f"material_request_{mr.id}.xlsx"
    file_path = os.path.join(UPLOAD_DIR, file_name); wb.save(file_path)
    return FileResponse(file_path, filename=f"{mr.name}.xlsx",
                       media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# =====================
# AI CHAT API
# =====================

@app.post("/api/chat/")
async def ai_chat(data: dict):
    """Чат с AI-помощником (заглушка, пока Qwen подключается к БД)."""
    user_msg = data.get("message", "")
    
    # TODO: Подключить OpenRouter + Function Calling для запросов к БД
    reply = (
        f"🤖 **Получено:** {user_msg}\n\n"
        "Сейчас я работаю в тестовом режиме. Скоро я смогу:\n"
        "• Отвечать на вопросы по объектам и сметам\n"
        "• Формировать отчёты по финансам\n"
        "• Помогать с расчётами материалов\n\n"
        "⏳ Подключение к базе знаний в процессе..."
    )
    return {"reply": reply}



# =====================
# ЗАЯВКИ НА РАБОТЫ
# =====================

@app.post("/work-batches/")
async def create_work_batch(data: schemas.WorkBatchCreate, db: Session = Depends(database.get_db)):
    obj = db.query(models.Object).filter(models.Object.id == data.object_id).first()
    if not obj: raise HTTPException(404, "Объект не найден")
    batch = models.WorkBatch(
        object_id=data.object_id, 
        name=data.name or f"Заявка: {obj.name}", 
        status="scheduled" if data.scheduled_at else "draft", 
        scheduled_at=data.scheduled_at, 
        notes=data.notes or "", 
        created_at=now_str()
    )
    db.add(batch); db.flush()
    for it in data.items:
        db.add(models.WorkBatchItem(batch_id=batch.id, estimate_item_id=it.estimate_item_id, name=it.name, unit=it.unit, quantity=it.quantity))
    for cid in data.contractor_ids:
        db.add(models.WorkBatchContractor(batch_id=batch.id, contractor_id=cid, status="pending"))
    db.commit(); db.refresh(batch)
    return {"id": batch.id, "name": batch.name, "status": batch.status, "items_count": len(data.items), "contractors_count": len(data.contractor_ids)}

@app.get("/work-batches/")
async def list_work_batches(object_id: Optional[int] = None, status: Optional[str] = None, db: Session = Depends(database.get_db)):
    q = db.query(models.WorkBatch).options(
        joinedload(models.WorkBatch.object), 
        joinedload(models.WorkBatch.items), 
        joinedload(models.WorkBatch.contractors).joinedload(models.WorkBatchContractor.contractor)
    )
    if object_id: q = q.filter(models.WorkBatch.object_id == object_id)
    if status: q = q.filter(models.WorkBatch.status == status)
    result = []
    for b in q.order_by(models.WorkBatch.id.desc()).all():
        result.append({
            "id": b.id, "object_id": b.object_id, "object_name": b.object.name if b.object else "", 
            "name": b.name, "status": b.status, "scheduled_at": b.scheduled_at, "created_at": b.created_at, "notes": b.notes, 
            "items": [{"id": i.id, "name": i.name, "unit": i.unit, "quantity": i.quantity} for i in b.items], 
            "contractors": [{"id": c.id, "contractor_id": c.contractor_id, "contractor_name": c.contractor.name if c.contractor else "", "status": c.status} for c in b.contractors]
        })
    return result

@app.get("/work-batches/{batch_id}")
async def get_work_batch(batch_id: int, db: Session = Depends(database.get_db)):
    b = db.query(models.WorkBatch).options(
        joinedload(models.WorkBatch.object), 
        joinedload(models.WorkBatch.items), 
        joinedload(models.WorkBatch.contractors).joinedload(models.WorkBatchContractor.contractor)
    ).filter(models.WorkBatch.id == batch_id).first()
    if not b: raise HTTPException(404)
    return {"id": b.id, "object_id": b.object_id, "object_name": b.object.name if b.object else "", "name": b.name, "status": b.status, "scheduled_at": b.scheduled_at, "notes": b.notes, "created_at": b.created_at, "items": [{"id": i.id, "name": i.name, "unit": i.unit, "quantity": i.quantity} for i in b.items], "contractors": [{"id": c.id, "contractor_id": c.contractor_id, "contractor_name": c.contractor.name if c.contractor else "", "status": c.status} for c in b.contractors]}

@app.put("/work-batches/{batch_id}")
async def update_work_batch(batch_id: int, data: dict, db: Session = Depends(database.get_db)):
    b = db.query(models.WorkBatch).filter(models.WorkBatch.id == batch_id).first()
    if not b: raise HTTPException(404)
    if "status" in data: b.status = data["status"]
    db.commit()
    return {"id": b.id, "status": b.status}

@app.delete("/work-batches/{batch_id}")
async def delete_work_batch(batch_id: int, db: Session = Depends(database.get_db)):
    b = db.query(models.WorkBatch).filter(models.WorkBatch.id == batch_id).first()
    if not b: raise HTTPException(404)
    db.delete(b); db.commit()
    return {"ok": True}

@app.post("/telegram-users/")
async def create_telegram_user(data: schemas.TelegramUserCreate, db: Session = Depends(database.get_db)):
    u = models.TelegramUser(contractor_id=data.contractor_id, telegram_id=data.telegram_id, telegram_username=data.telegram_username or "", first_name=data.first_name or "", last_name=data.last_name or "", created_at=now_str())
    db.add(u); db.commit(); db.refresh(u)
    return {"id": u.id, "telegram_id": u.telegram_id}

@app.get("/telegram-users/")
async def list_telegram_users(contractor_id: Optional[int] = None, db: Session = Depends(database.get_db)):
    q = db.query(models.TelegramUser).options(joinedload(models.TelegramUser.contractor))
    if contractor_id: q = q.filter(models.TelegramUser.contractor_id == contractor_id)
    return [{"id": u.id, "contractor_id": u.contractor_id, "contractor_name": u.contractor.name if u.contractor else "", "telegram_id": u.telegram_id} for u in q.all()]

# =====================
# ГЛАВНАЯ
# =====================

@app.get("/")
async def root():
    idx = FRONTEND_DIR / "index.html"
    if idx.exists(): return FileResponse(str(idx))
    return {"message": "CRM v4.0"}